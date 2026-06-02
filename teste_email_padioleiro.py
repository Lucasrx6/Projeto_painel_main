"""
Teste de envio do relatório diário de movimentações do Padioleiro.
Consulta dados reais do banco e envia email com Excel anexado.
"""

import os, sys, smtplib, tempfile
from datetime import datetime, date
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

import psycopg2
from psycopg2.extras import RealDictCursor
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import ColorScaleRule

_BASE_DIR = os.path.dirname(os.path.abspath(__file__))
load_dotenv(os.path.join(_BASE_DIR, '.env'))

DESTINATARIO = 'lucas.oliveira@saofranciscodf.med.br'

_HAC      = "9B1C24"
_VERDE    = "28A745"
_VERMELHO = "DC3545"
_LARANJA  = "E67E00"
_AZUL     = "17A2B8"
_BRANCO   = "FFFFFF"
_ZEBRA    = "FEF0F0"
_FUNDO_H  = "F5D0D3"


# ── BANCO ─────────────────────────────────────────────────────

def _get_conn():
    return psycopg2.connect(
        host=os.getenv('DB_HOST', 'localhost'),
        dbname=os.getenv('DB_NAME', 'postgres'),
        user=os.getenv('DB_USER', 'postgres'),
        password=os.getenv('DB_PASSWORD', 'postgres'),
        port=int(os.getenv('DB_PORT', '5432')),
    )


def buscar_dados():
    conn = _get_conn()
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:

            cur.execute("""
                SELECT
                    COUNT(*) AS total,
                    COUNT(*) FILTER (WHERE status = 'concluido')  AS concluidos,
                    COUNT(*) FILTER (WHERE status = 'cancelado')  AS cancelados,
                    COUNT(*) FILTER (WHERE status IN ('aguardando','aceito','em_transporte')) AS em_aberto,
                    COUNT(*) FILTER (WHERE prioridade = 'urgente') AS urgentes,
                    ROUND(AVG(CASE WHEN dt_aceite IS NOT NULL
                        THEN EXTRACT(EPOCH FROM (dt_aceite - criado_em)) / 60 END)::numeric, 1) AS media_aceite_min,
                    ROUND(AVG(CASE WHEN dt_conclusao IS NOT NULL
                        THEN EXTRACT(EPOCH FROM (dt_conclusao - criado_em)) / 60 END)::numeric, 1) AS media_total_min
                FROM padioleiro_chamados WHERE DATE(criado_em) = CURRENT_DATE
            """)
            resumo = dict(cur.fetchone() or {})

            cur.execute("""
                SELECT
                    COALESCE(padioleiro_nome,'(não atribuído)') AS padioleiro,
                    COUNT(*) AS total,
                    COUNT(*) FILTER (WHERE status='concluido')  AS concluidos,
                    COUNT(*) FILTER (WHERE status='cancelado')  AS cancelados,
                    COUNT(*) FILTER (WHERE prioridade='urgente') AS urgentes,
                    ROUND(AVG(CASE WHEN dt_aceite IS NOT NULL
                        THEN EXTRACT(EPOCH FROM (dt_aceite - criado_em))/60 END)::numeric,1) AS media_aceite_min,
                    ROUND(AVG(CASE WHEN dt_inicio_transporte IS NOT NULL AND dt_aceite IS NOT NULL
                        THEN EXTRACT(EPOCH FROM (dt_inicio_transporte - dt_aceite))/60 END)::numeric,1) AS media_deslocamento_min,
                    ROUND(AVG(CASE WHEN dt_conclusao IS NOT NULL AND dt_inicio_transporte IS NOT NULL
                        THEN EXTRACT(EPOCH FROM (dt_conclusao - dt_inicio_transporte))/60 END)::numeric,1) AS media_transporte_min,
                    ROUND(AVG(CASE WHEN dt_conclusao IS NOT NULL
                        THEN EXTRACT(EPOCH FROM (dt_conclusao - criado_em))/60 END)::numeric,1) AS media_total_min
                FROM padioleiro_chamados WHERE DATE(criado_em) = CURRENT_DATE
                GROUP BY padioleiro_nome ORDER BY total DESC
            """)
            por_padioleiro = [dict(r) for r in cur.fetchall()]

            cur.execute("""
                SELECT
                    COALESCE(setor_origem_nome,'(sem setor)') AS setor,
                    COUNT(*) AS total,
                    COUNT(*) FILTER (WHERE status='concluido')  AS concluidos,
                    COUNT(*) FILTER (WHERE status='cancelado')  AS cancelados,
                    COUNT(*) FILTER (WHERE prioridade='urgente') AS urgentes
                FROM padioleiro_chamados WHERE DATE(criado_em) = CURRENT_DATE
                GROUP BY setor_origem_nome ORDER BY total DESC LIMIT 20
            """)
            por_setor = [dict(r) for r in cur.fetchall()]

            cur.execute("""
                SELECT
                    EXTRACT(HOUR FROM criado_em)::int AS hora,
                    COUNT(*) AS total,
                    COUNT(*) FILTER (WHERE status='concluido')   AS concluidos,
                    COUNT(*) FILTER (WHERE prioridade='urgente') AS urgentes
                FROM padioleiro_chamados WHERE DATE(criado_em) = CURRENT_DATE
                GROUP BY 1 ORDER BY 1
            """)
            por_hora = [dict(r) for r in cur.fetchall()]

            cur.execute("""
                SELECT
                    DATE(criado_em) AS data,
                    COUNT(*) AS total,
                    COUNT(*) FILTER (WHERE status='concluido') AS concluidos,
                    COUNT(*) FILTER (WHERE status='cancelado') AS cancelados,
                    ROUND(AVG(CASE WHEN dt_conclusao IS NOT NULL
                        THEN EXTRACT(EPOCH FROM (dt_conclusao - criado_em))/60 END)::numeric,1) AS media_total_min
                FROM padioleiro_chamados
                WHERE criado_em >= CURRENT_DATE - INTERVAL '6 days'
                GROUP BY DATE(criado_em) ORDER BY 1
            """)
            tendencia_7d = [dict(r) for r in cur.fetchall()]

            cur.execute("""
                SELECT
                    id, COALESCE(tipo_movimento_nome,'--') AS tipo,
                    nm_paciente, leito_origem, setor_origem_nome, destino_nome,
                    prioridade, status, solicitante_nome,
                    COALESCE(padioleiro_nome,'--') AS padioleiro, observacao,
                    TO_CHAR(criado_em,'HH24:MI')             AS criado_em,
                    TO_CHAR(dt_aceite,'HH24:MI')             AS dt_aceite,
                    TO_CHAR(dt_inicio_transporte,'HH24:MI')  AS dt_inicio_transporte,
                    TO_CHAR(dt_conclusao,'HH24:MI')          AS dt_conclusao,
                    TO_CHAR(dt_cancelamento,'HH24:MI')       AS dt_cancelamento,
                    motivo_cancelamento,
                    CASE WHEN dt_aceite IS NOT NULL
                         THEN ROUND(EXTRACT(EPOCH FROM (dt_aceite - criado_em))/60)::int END AS t_aceite_min,
                    CASE WHEN dt_inicio_transporte IS NOT NULL AND dt_aceite IS NOT NULL
                         THEN ROUND(EXTRACT(EPOCH FROM (dt_inicio_transporte - dt_aceite))/60)::int END AS t_deslocamento_min,
                    CASE WHEN dt_conclusao IS NOT NULL AND dt_inicio_transporte IS NOT NULL
                         THEN ROUND(EXTRACT(EPOCH FROM (dt_conclusao - dt_inicio_transporte))/60)::int END AS t_transporte_min,
                    CASE WHEN dt_conclusao IS NOT NULL
                         THEN ROUND(EXTRACT(EPOCH FROM (dt_conclusao - criado_em))/60)::int
                         WHEN status='cancelado' AND dt_cancelamento IS NOT NULL
                         THEN ROUND(EXTRACT(EPOCH FROM (dt_cancelamento - criado_em))/60)::int END AS t_total_min
                FROM padioleiro_chamados WHERE DATE(criado_em) = CURRENT_DATE
                ORDER BY criado_em DESC
            """)
            todos_hoje = [dict(r) for r in cur.fetchall()]

            cur.execute("""
                SELECT id, nm_paciente, setor_origem_nome, destino_nome,
                    padioleiro_nome, solicitante_nome,
                    TO_CHAR(criado_em,'HH24:MI')       AS hora_criacao,
                    TO_CHAR(dt_cancelamento,'HH24:MI') AS hora_cancelamento,
                    motivo_cancelamento
                FROM padioleiro_chamados
                WHERE DATE(criado_em) = CURRENT_DATE AND status = 'cancelado'
                ORDER BY dt_cancelamento DESC
            """)
            cancelados = [dict(r) for r in cur.fetchall()]

        return resumo, por_padioleiro, por_setor, por_hora, tendencia_7d, todos_hoje, cancelados
    finally:
        conn.close()


# ── EXCEL HELPERS ─────────────────────────────────────────────

def _hdr(cell, cor=_HAC):
    cell.font = Font(bold=True, color=_BRANCO, size=11)
    cell.fill = PatternFill("solid", fgColor=cor)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _borda(cell)

def _borda(cell):
    s = Side(style="thin", color="CCCCCC")
    cell.border = Border(left=s, right=s, top=s, bottom=s)

def _autowidth(ws, min_w=10, max_w=45):
    for col in ws.columns:
        letra = get_column_letter(col[0].column)
        w = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[letra].width = min(max(w + 3, min_w), max_w)

def _titulo(ws, texto, ncols, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=texto)
    c.font = Font(bold=True, color=_BRANCO, size=13)
    c.fill = PatternFill("solid", fgColor=_HAC)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28

def _serie_cor(serie, cor):
    serie.graphicalProperties.solidFill = cor

def _cor_tempo(ws, col_letra, row_ini, row_fim):
    rng = f"{col_letra}{row_ini}:{col_letra}{row_fim}"
    ws.conditional_formatting.add(rng, ColorScaleRule(
        start_type='num', start_value=0,  start_color="63BE7B",
        mid_type='num',   mid_value=20,   mid_color="FFEB84",
        end_type='num',   end_value=60,   end_color="F8696B",
    ))


# ── ABAs DO EXCEL ─────────────────────────────────────────────

def _aba_resumo(wb, resumo, now):
    ws = wb.create_sheet("Resumo do Dia")
    ws.sheet_view.showGridLines = False
    _titulo(ws, f"MOVIMENTAÇÕES DO PADIOLEIRO — HAC — {now.strftime('%d/%m/%Y %H:%M')}", 5)

    total      = int(resumo.get('total') or 0)
    concluidos = int(resumo.get('concluidos') or 0)
    cancelados = int(resumo.get('cancelados') or 0)
    em_aberto  = int(resumo.get('em_aberto') or 0)
    urgentes   = int(resumo.get('urgentes') or 0)
    taxa       = round(concluidos / total * 100) if total else 0

    kpis = [
        ("Indicador", "Valor", None),
        ("Total de Chamados",           total,                                _HAC),
        ("Concluídos",                  concluidos,                           _VERDE),
        ("Cancelados",                  cancelados,                           _VERMELHO),
        ("Em Aberto",                   em_aberto,                            _AZUL),
        ("Urgentes",                    urgentes,                             _LARANJA),
        ("Taxa de Conclusão (%)",       taxa,                                 _HAC),
        ("Tempo Médio até Aceite (min)", resumo.get('media_aceite_min') or '--', None),
        ("Tempo Médio Total (min)",     resumo.get('media_total_min') or '--',   None),
    ]
    for i, (ind, val, cor) in enumerate(kpis, 3):
        c1 = ws.cell(row=i, column=1, value=ind)
        c2 = ws.cell(row=i, column=2, value=val)
        if i == 3:
            _hdr(c1); _hdr(c2)
        else:
            c1.font = Font(bold=True)
            c1.fill = PatternFill("solid", fgColor=_FUNDO_H)
            _borda(c1); _borda(c2)
            if cor:
                c2.font = Font(bold=True, color=_BRANCO)
                c2.fill = PatternFill("solid", fgColor=cor)

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18

    # Tabela auxiliar para gráfico de pizza (coluna D/E)
    pie_items = [("Status", "Qtd"), ("Concluídos", concluidos), ("Cancelados", cancelados), ("Em Aberto", em_aberto)]
    for i, (s, v) in enumerate(pie_items, 3):
        ws.cell(row=i, column=4, value=s).font = Font(bold=(i == 3))
        ws.cell(row=i, column=5, value=v)
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 10

    if total > 0:
        from openpyxl.chart import PieChart
        pie = PieChart()
        pie.title = "Distribuição por Status"
        pie.style = 10; pie.height = 12; pie.width = 16
        data   = Reference(ws, min_col=5, max_col=5, min_row=3, max_row=6)
        labels = Reference(ws, min_col=4, min_row=4, max_row=6)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        for idx, cor in enumerate([_VERDE, _VERMELHO, _AZUL]):
            pt = DataPoint(idx=idx)
            pt.graphicalProperties.solidFill = cor
            pie.series[0].dPt.append(pt)
        ws.add_chart(pie, "D6")


def _aba_por_padioleiro(wb, por_padioleiro):
    ws = wb.create_sheet("Por Padioleiro")
    ws.sheet_view.showGridLines = False

    # Tabela de volumes
    hv = ["Padioleiro", "Total", "Concluídos", "Cancelados", "Urgentes"]
    _titulo(ws, "VOLUMES POR PADIOLEIRO", len(hv), 1)
    for j, h in enumerate(hv, 1):
        _hdr(ws.cell(row=2, column=j, value=h))

    for i, p in enumerate(por_padioleiro, 3):
        zebra = i % 2 == 0
        for j, (key, cor_txt) in enumerate([
            ('padioleiro', None), ('total', None), ('concluidos', _VERDE),
            ('cancelados', _VERMELHO), ('urgentes', _LARANJA)
        ], 1):
            cell = ws.cell(row=i, column=j, value=p.get(key))
            _borda(cell)
            if zebra: cell.fill = PatternFill("solid", fgColor=_ZEBRA)
            if j == 1: cell.font = Font(bold=True)
            elif cor_txt: cell.font = Font(bold=True, color=cor_txt)

    last_vol = 2 + len(por_padioleiro)

    if por_padioleiro:
        c = BarChart()
        c.type = "bar"; c.grouping = "clustered"
        c.title = "Movimentos por Padioleiro"
        c.style = 10; c.height = 12; c.width = 22
        c.x_axis.title = "Quantidade"
        c.add_data(Reference(ws, min_col=2, max_col=5, min_row=2, max_row=last_vol), titles_from_data=True)
        c.set_categories(Reference(ws, min_col=1, min_row=3, max_row=last_vol))
        for idx, cor in enumerate([_HAC, _VERDE, _VERMELHO, _LARANJA]):
            if idx < len(c.series): _serie_cor(c.series[idx], cor)
        ws.add_chart(c, f"A{last_vol + 3}")
        prox = last_vol + 20
    else:
        prox = last_vol + 3

    # Tabela de tempos
    ht = ["Padioleiro", "T.Aceite(min)", "T.Desloc.(min)", "T.Transp.(min)", "T.Total(min)"]
    _titulo(ws, "TEMPOS MÉDIOS POR PADIOLEIRO", len(ht), prox)
    for j, h in enumerate(ht, 1):
        _hdr(ws.cell(row=prox + 1, column=j, value=h))

    for i, p in enumerate(por_padioleiro, prox + 2):
        zebra = i % 2 == 0
        keys = ['padioleiro','media_aceite_min','media_deslocamento_min','media_transporte_min','media_total_min']
        for j, key in enumerate(keys, 1):
            v = p.get(key)
            cell = ws.cell(row=i, column=j, value=float(v) if v is not None and j > 1 else v)
            _borda(cell)
            if zebra: cell.fill = PatternFill("solid", fgColor=_ZEBRA)
            if j == 1: cell.font = Font(bold=True)

    last_tem = prox + 1 + len(por_padioleiro)
    if por_padioleiro:
        for col_l in ['B', 'C', 'D', 'E']:
            _cor_tempo(ws, col_l, prox + 2, last_tem)
        c2 = BarChart()
        c2.type = "bar"; c2.grouping = "clustered"
        c2.title = "Tempos Médios por Padioleiro (min)"
        c2.style = 10; c2.height = 12; c2.width = 22
        c2.x_axis.title = "Minutos"
        c2.add_data(Reference(ws, min_col=2, max_col=5, min_row=prox+1, max_row=last_tem), titles_from_data=True)
        c2.set_categories(Reference(ws, min_col=1, min_row=prox+2, max_row=last_tem))
        for idx, cor in enumerate([_AZUL, _HAC, _VERDE, _LARANJA]):
            if idx < len(c2.series): _serie_cor(c2.series[idx], cor)
        ws.add_chart(c2, f"A{last_tem + 3}")

    _autowidth(ws)


def _aba_por_setor(wb, por_setor):
    ws = wb.create_sheet("Por Setor")
    ws.sheet_view.showGridLines = False
    headers = ["Setor", "Total", "Concluídos", "Cancelados", "Urgentes"]
    _titulo(ws, "CHAMADOS POR SETOR DE ORIGEM", len(headers))
    for j, h in enumerate(headers, 1):
        _hdr(ws.cell(row=2, column=j, value=h))

    for i, s in enumerate(por_setor, 3):
        zebra = i % 2 == 0
        for j, (key, cor_txt) in enumerate([
            ('setor',None),('total',None),('concluidos',_VERDE),('cancelados',_VERMELHO),('urgentes',_LARANJA)
        ], 1):
            cell = ws.cell(row=i, column=j, value=s.get(key))
            _borda(cell)
            if zebra: cell.fill = PatternFill("solid", fgColor=_ZEBRA)
            if j == 1: cell.font = Font(bold=True)
            elif cor_txt: cell.font = Font(bold=True, color=cor_txt)

    last_row = 2 + len(por_setor)
    if por_setor:
        c = BarChart()
        c.type = "bar"; c.grouping = "clustered"
        c.title = "Chamados por Setor"
        c.style = 10; c.height = max(12, len(por_setor) * 0.9); c.width = 22
        c.x_axis.title = "Quantidade"
        c.add_data(Reference(ws, min_col=2, max_col=5, min_row=2, max_row=last_row), titles_from_data=True)
        c.set_categories(Reference(ws, min_col=1, min_row=3, max_row=last_row))
        for idx, cor in enumerate([_HAC, _VERDE, _VERMELHO, _LARANJA]):
            if idx < len(c.series): _serie_cor(c.series[idx], cor)
        ws.add_chart(c, f"A{last_row + 3}")
    _autowidth(ws)


def _aba_por_hora(wb, por_hora):
    ws = wb.create_sheet("Distribuição por Hora")
    ws.sheet_view.showGridLines = False
    headers = ["Hora", "Total", "Concluídos", "Urgentes"]
    _titulo(ws, "CHAMADOS POR HORA DO DIA", len(headers))
    for j, h in enumerate(headers, 1):
        _hdr(ws.cell(row=2, column=j, value=h))

    hora_dict = {r['hora']: r for r in por_hora}
    for hora in range(24):
        r = hora_dict.get(hora, {})
        row_n = hora + 3
        zebra = hora % 2 == 0
        for j, v in enumerate([f"{hora:02d}h", int(r.get('total',0)), int(r.get('concluidos',0)), int(r.get('urgentes',0))], 1):
            cell = ws.cell(row=row_n, column=j, value=v)
            _borda(cell)
            if zebra: cell.fill = PatternFill("solid", fgColor=_ZEBRA)

    c = BarChart()
    c.type = "col"; c.grouping = "clustered"
    c.title = "Distribuição de Chamados por Hora"
    c.style = 10; c.height = 14; c.width = 26
    c.y_axis.title = "Chamados"; c.x_axis.title = "Hora"
    c.add_data(Reference(ws, min_col=2, max_col=4, min_row=2, max_row=26), titles_from_data=True)
    c.set_categories(Reference(ws, min_col=1, min_row=3, max_row=26))
    for idx, cor in enumerate([_HAC, _VERDE, _LARANJA]):
        if idx < len(c.series): _serie_cor(c.series[idx], cor)
    ws.add_chart(c, "F2")
    _autowidth(ws)


def _aba_tendencia(wb, tendencia_7d):
    ws = wb.create_sheet("Tendência 7 Dias")
    ws.sheet_view.showGridLines = False
    headers = ["Data", "Total", "Concluídos", "Cancelados", "T.Médio(min)"]
    _titulo(ws, "TENDÊNCIA — ÚLTIMOS 7 DIAS", len(headers))
    for j, h in enumerate(headers, 1):
        _hdr(ws.cell(row=2, column=j, value=h))

    for i, d in enumerate(tendencia_7d, 3):
        zebra = i % 2 == 0
        dt = d.get('data')
        data_fmt = dt.strftime('%d/%m') if hasattr(dt, 'strftime') else str(dt or '')
        vals = [data_fmt, d.get('total'), d.get('concluidos'), d.get('cancelados'), d.get('media_total_min')]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=j, value=float(v) if v is not None and j == 5 else v)
            _borda(cell)
            if zebra: cell.fill = PatternFill("solid", fgColor=_ZEBRA)

    last_row = 2 + len(tendencia_7d)
    if tendencia_7d:
        c = LineChart()
        c.title = "Tendência de Chamados — Últimos 7 Dias"
        c.style = 10; c.height = 14; c.width = 22
        c.y_axis.title = "Chamados"; c.x_axis.title = "Data"
        c.add_data(Reference(ws, min_col=2, max_col=4, min_row=2, max_row=last_row), titles_from_data=True)
        c.set_categories(Reference(ws, min_col=1, min_row=3, max_row=last_row))
        for idx, cor in enumerate([_HAC, _VERDE, _VERMELHO]):
            if idx < len(c.series):
                c.series[idx].graphicalProperties.line.solidFill = cor
        ws.add_chart(c, f"A{last_row + 3}")
    _autowidth(ws)


def _aba_todos_chamados(wb, todos_hoje, now):
    ws = wb.create_sheet("Todos os Chamados")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    headers = [
        "#", "Tipo", "Paciente", "Leito", "Setor Origem", "Destino",
        "Prioridade", "Status", "Solicitante", "Padioleiro",
        "Criado", "Aceito", "Ini.Transp.", "Conclusão", "Cancelado",
        "T.Aceite(min)", "T.Desloc.(min)", "T.Transp.(min)", "T.Total(min)",
        "Motivo Cancelamento", "Obs."
    ]
    _titulo(ws, f"TODOS OS CHAMADOS — {now.strftime('%d/%m/%Y')}", len(headers))
    for j, h in enumerate(headers, 1):
        _hdr(ws.cell(row=2, column=j, value=h))
    ws.row_dimensions[2].height = 22

    _STATUS_COR = {
        'concluido': _VERDE, 'cancelado': _VERMELHO,
        'aguardando': _HAC, 'aceito': _AZUL, 'em_transporte': _LARANJA
    }
    keys = [
        'id','tipo','nm_paciente','leito_origem','setor_origem_nome','destino_nome',
        'prioridade','status','solicitante_nome','padioleiro',
        'criado_em','dt_aceite','dt_inicio_transporte','dt_conclusao','dt_cancelamento',
        't_aceite_min','t_deslocamento_min','t_transporte_min','t_total_min',
        'motivo_cancelamento','observacao'
    ]
    for i, chamado in enumerate(todos_hoje, 3):
        zebra = i % 2 == 0
        for j, key in enumerate(keys, 1):
            v = chamado.get(key)
            cell = ws.cell(row=i, column=j, value=v)
            _borda(cell)
            cell.alignment = Alignment(vertical="center", wrap_text=(j >= 20))
            if zebra and key not in ('prioridade', 'status'):
                cell.fill = PatternFill("solid", fgColor=_ZEBRA)
            if key == 'status' and v in _STATUS_COR:
                cell.font = Font(bold=True, color=_BRANCO)
                cell.fill = PatternFill("solid", fgColor=_STATUS_COR[v])
            elif key == 'prioridade' and v == 'urgente':
                cell.font = Font(bold=True, color=_BRANCO)
                cell.fill = PatternFill("solid", fgColor=_LARANJA)

    last_row = 2 + len(todos_hoje)
    if todos_hoje:
        for col_n in range(16, 20):
            _cor_tempo(ws, get_column_letter(col_n), 3, last_row)
    _autowidth(ws)


def _aba_cancelamentos(wb, cancelados):
    ws = wb.create_sheet("Cancelamentos")
    ws.sheet_view.showGridLines = False
    headers = ["#", "Paciente", "Setor Origem", "Destino", "Padioleiro",
               "Solicitante", "H.Criação", "H.Cancel.", "Motivo"]
    _titulo(ws, "CANCELAMENTOS DO DIA", len(headers))
    for j, h in enumerate(headers, 1):
        _hdr(ws.cell(row=2, column=j, value=h), cor=_VERMELHO)

    keys = ['id','nm_paciente','setor_origem_nome','destino_nome','padioleiro_nome',
            'solicitante_nome','hora_criacao','hora_cancelamento','motivo_cancelamento']
    for i, c in enumerate(cancelados, 3):
        zebra = i % 2 == 0
        for j, key in enumerate(keys, 1):
            cell = ws.cell(row=i, column=j, value=c.get(key))
            _borda(cell)
            if zebra: cell.fill = PatternFill("solid", fgColor="FEE2E2")
            if j == 9:
                cell.alignment = Alignment(wrap_text=True)
                cell.font = Font(color=_VERMELHO)
    _autowidth(ws)


def gerar_excel(resumo, por_padioleiro, por_setor, por_hora, tendencia_7d, todos_hoje, cancelados, now):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _aba_resumo(wb, resumo, now)
    _aba_por_padioleiro(wb, por_padioleiro)
    _aba_por_setor(wb, por_setor)
    _aba_por_hora(wb, por_hora)
    _aba_tendencia(wb, tendencia_7d)
    _aba_todos_chamados(wb, todos_hoje, now)
    if cancelados:
        _aba_cancelamentos(wb, cancelados)
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(tmp.name)
    tmp.close()
    return tmp.name


# ── HTML ──────────────────────────────────────────────────────

def _fmt(val, sufixo=''):
    if val is None:
        return '<span style="color:#aaa;">--</span>'
    return f'{val}{sufixo}'


def gerar_html(resumo, por_padioleiro, por_setor, cancelados, now):
    total        = resumo.get('total', 0) or 0
    concluidos   = resumo.get('concluidos', 0) or 0
    cancelados_n = resumo.get('cancelados', 0) or 0
    em_aberto    = resumo.get('em_aberto', 0) or 0
    urgentes     = resumo.get('urgentes', 0) or 0
    media_aceite = resumo.get('media_aceite_min')
    media_total  = resumo.get('media_total_min')
    taxa_conclusao = f'{round(concluidos / total * 100)}%' if total else '--'

    linhas_pad = ''
    for p in por_padioleiro:
        linhas_pad += (
            '<tr>'
            f'<td style="padding:8px 12px;border-bottom:1px solid #eee;font-weight:600;">{p["padioleiro"]}</td>'
            f'<td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:center;">{p["total"]}</td>'
            f'<td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:center;color:#28a745;font-weight:700;">{p["concluidos"]}</td>'
            f'<td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:center;color:#dc3545;">{p["cancelados"]}</td>'
            f'<td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:center;color:#e67e00;font-weight:700;">{p["urgentes"]}</td>'
            f'<td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:center;">{_fmt(p.get("media_aceite_min"), " min")}</td>'
            f'<td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:center;">{_fmt(p.get("media_transporte_min"), " min")}</td>'
            f'<td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:center;">{_fmt(p.get("media_total_min"), " min")}</td>'
            '</tr>'
        )
    if not linhas_pad:
        linhas_pad = '<tr><td colspan="8" style="padding:16px;text-align:center;color:#aaa;">Nenhum dado hoje</td></tr>'

    linhas_set = ''
    for s in por_setor:
        linhas_set += (
            '<tr>'
            f'<td style="padding:8px 12px;border-bottom:1px solid #eee;font-weight:600;">{s["setor"]}</td>'
            f'<td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:center;">{s["total"]}</td>'
            f'<td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:center;color:#28a745;font-weight:700;">{s["concluidos"]}</td>'
            f'<td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:center;color:#dc3545;">{s["cancelados"]}</td>'
            f'<td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:center;color:#e67e00;font-weight:700;">{s["urgentes"]}</td>'
            '</tr>'
        )
    if not linhas_set:
        linhas_set = '<tr><td colspan="5" style="padding:16px;text-align:center;color:#aaa;">Nenhum dado hoje</td></tr>'

    bloco_cancelados = ''
    if cancelados:
        linhas_canc = ''
        for c in cancelados:
            linhas_canc += (
                '<tr>'
                f'<td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:center;color:#aaa;">{c.get("hora_criacao","--")} → {c.get("hora_cancelamento","--")}</td>'
                f'<td style="padding:8px 12px;border-bottom:1px solid #eee;">{c.get("nm_paciente","--")}</td>'
                f'<td style="padding:8px 12px;border-bottom:1px solid #eee;">{c.get("setor_origem_nome","--")}</td>'
                f'<td style="padding:8px 12px;border-bottom:1px solid #eee;">{c.get("destino_nome","--")}</td>'
                f'<td style="padding:8px 12px;border-bottom:1px solid #eee;">{c.get("padioleiro_nome","--")}</td>'
                f'<td style="padding:8px 12px;border-bottom:1px solid #eee;color:#dc3545;">{c.get("motivo_cancelamento","--")}</td>'
                '</tr>'
            )
        bloco_cancelados = f"""
        <h3 style="color:#9B1C24;font-size:16px;margin:32px 0 12px;">Cancelamentos do Dia ({len(cancelados)})</h3>
        <table style="width:100%;border-collapse:collapse;border-radius:8px;overflow:hidden;box-shadow:0 1px 4px rgba(0,0,0,.08);">
          <thead><tr style="background:#9B1C24;">
            <th style="padding:10px 12px;color:#fff;font-size:13px;">Horário</th>
            <th style="padding:10px 12px;color:#fff;text-align:left;font-size:13px;">Paciente</th>
            <th style="padding:10px 12px;color:#fff;text-align:left;font-size:13px;">Origem</th>
            <th style="padding:10px 12px;color:#fff;text-align:left;font-size:13px;">Destino</th>
            <th style="padding:10px 12px;color:#fff;text-align:left;font-size:13px;">Padioleiro</th>
            <th style="padding:10px 12px;color:#fff;text-align:left;font-size:13px;">Motivo</th>
          </tr></thead>
          <tbody>{linhas_canc}</tbody>
        </table>"""

    return f"""<!DOCTYPE html>
<html lang="pt-BR"><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#f4f4f4;font-family:Arial,sans-serif;">
<div style="max-width:720px;margin:24px auto;">

  <div style="background:#9B1C24;padding:28px 32px;border-radius:8px 8px 0 0;">
    <div style="color:#fff;font-size:11px;letter-spacing:1px;text-transform:uppercase;opacity:.8;">Hospital Anchieta Ceilândia</div>
    <div style="color:#fff;font-size:22px;font-weight:700;margin-top:4px;">Movimentações do Padioleiro</div>
    <div style="color:#fff;font-size:13px;opacity:.85;margin-top:4px;">{now.strftime('%A, %d de %B de %Y — %H:%M').capitalize()}</div>
  </div>

  <div style="background:#fff;padding:24px 32px;">
    <table style="width:100%;border-collapse:collapse;">
      <tr>
        <td style="width:16%;padding:0 8px 0 0;">
          <div style="background:#f8f9fa;border-radius:8px;padding:16px;text-align:center;border-left:4px solid #9B1C24;">
            <div style="font-size:28px;font-weight:700;color:#9B1C24;">{total}</div>
            <div style="font-size:11px;color:#666;margin-top:4px;">TOTAL</div>
          </div>
        </td>
        <td style="width:16%;padding:0 8px;">
          <div style="background:#f8f9fa;border-radius:8px;padding:16px;text-align:center;border-left:4px solid #28a745;">
            <div style="font-size:28px;font-weight:700;color:#28a745;">{concluidos}</div>
            <div style="font-size:11px;color:#666;margin-top:4px;">CONCLUÍDOS</div>
          </div>
        </td>
        <td style="width:16%;padding:0 8px;">
          <div style="background:#f8f9fa;border-radius:8px;padding:16px;text-align:center;border-left:4px solid #dc3545;">
            <div style="font-size:28px;font-weight:700;color:#dc3545;">{cancelados_n}</div>
            <div style="font-size:11px;color:#666;margin-top:4px;">CANCELADOS</div>
          </div>
        </td>
        <td style="width:16%;padding:0 8px;">
          <div style="background:#f8f9fa;border-radius:8px;padding:16px;text-align:center;border-left:4px solid #17a2b8;">
            <div style="font-size:28px;font-weight:700;color:#17a2b8;">{em_aberto}</div>
            <div style="font-size:11px;color:#666;margin-top:4px;">EM ABERTO</div>
          </div>
        </td>
        <td style="width:16%;padding:0 8px;">
          <div style="background:#f8f9fa;border-radius:8px;padding:16px;text-align:center;border-left:4px solid #e67e00;">
            <div style="font-size:28px;font-weight:700;color:#e67e00;">{urgentes}</div>
            <div style="font-size:11px;color:#666;margin-top:4px;">URGENTES</div>
          </div>
        </td>
        <td style="width:20%;padding:0 0 0 8px;">
          <div style="background:#f8f9fa;border-radius:8px;padding:16px;text-align:center;border-left:4px solid #6c757d;">
            <div style="font-size:28px;font-weight:700;color:#6c757d;">{taxa_conclusao}</div>
            <div style="font-size:11px;color:#666;margin-top:4px;">TAXA CONCLUSÃO</div>
          </div>
        </td>
      </tr>
    </table>
    <div style="margin-top:16px;background:#f8f9fa;border-radius:8px;padding:14px 20px;">
      <span style="font-size:13px;color:#444;">⏱ <strong>T.médio aceite:</strong> {_fmt(media_aceite, ' min')}</span>
      <span style="font-size:13px;color:#444;margin-left:32px;">⏱ <strong>T.médio total:</strong> {_fmt(media_total, ' min')}</span>
    </div>
  </div>

  <div style="background:#fff;padding:0 32px 24px;margin-top:2px;">
    <h3 style="color:#9B1C24;font-size:16px;margin:0 0 12px;padding-top:24px;">Desempenho por Padioleiro</h3>
    <table style="width:100%;border-collapse:collapse;border-radius:8px;overflow:hidden;box-shadow:0 1px 4px rgba(0,0,0,.08);">
      <thead><tr style="background:#9B1C24;">
        <th style="padding:10px 12px;color:#fff;text-align:left;font-size:13px;">Padioleiro</th>
        <th style="padding:10px 12px;color:#fff;text-align:center;font-size:13px;">Total</th>
        <th style="padding:10px 12px;color:#fff;text-align:center;font-size:13px;">Concluídos</th>
        <th style="padding:10px 12px;color:#fff;text-align:center;font-size:13px;">Cancelados</th>
        <th style="padding:10px 12px;color:#fff;text-align:center;font-size:13px;">Urgentes</th>
        <th style="padding:10px 12px;color:#fff;text-align:center;font-size:13px;">T.Aceite</th>
        <th style="padding:10px 12px;color:#fff;text-align:center;font-size:13px;">T.Transporte</th>
        <th style="padding:10px 12px;color:#fff;text-align:center;font-size:13px;">T.Total</th>
      </tr></thead>
      <tbody>{linhas_pad}</tbody>
    </table>
  </div>

  <div style="background:#fff;padding:0 32px 24px;margin-top:2px;">
    <h3 style="color:#9B1C24;font-size:16px;margin:0 0 12px;padding-top:24px;">Solicitações por Setor</h3>
    <table style="width:100%;border-collapse:collapse;border-radius:8px;overflow:hidden;box-shadow:0 1px 4px rgba(0,0,0,.08);">
      <thead><tr style="background:#9B1C24;">
        <th style="padding:10px 12px;color:#fff;text-align:left;font-size:13px;">Setor</th>
        <th style="padding:10px 12px;color:#fff;text-align:center;font-size:13px;">Total</th>
        <th style="padding:10px 12px;color:#fff;text-align:center;font-size:13px;">Concluídos</th>
        <th style="padding:10px 12px;color:#fff;text-align:center;font-size:13px;">Cancelados</th>
        <th style="padding:10px 12px;color:#fff;text-align:center;font-size:13px;">Urgentes</th>
      </tr></thead>
      <tbody>{linhas_set}</tbody>
    </table>
  </div>

  {'<div style="background:#fff;padding:0 32px 24px;margin-top:2px;">' + bloco_cancelados + '</div>' if bloco_cancelados else ''}

  <div style="background:#9B1C24;padding:16px 32px;border-radius:0 0 8px 8px;text-align:center;">
    <span style="color:rgba(255,255,255,.7);font-size:11px;">
      Central de Informações HAC — Gerado em {now.strftime('%d/%m/%Y às %H:%M')} — Painel 36
      &nbsp;|&nbsp; Excel em anexo com gráficos detalhados
    </span>
  </div>

</div></body></html>"""


# ── ENVIO ─────────────────────────────────────────────────────

def enviar(destinatario, html, excel_path, now):
    smtp_host = os.getenv('SMTP_HOST', '')
    smtp_port = int(os.getenv('SMTP_PORT', '587'))
    smtp_user = os.getenv('SMTP_USER', '')
    smtp_pass = os.getenv('SMTP_PASS', '')
    smtp_from = os.getenv('SMTP_FROM', '') or smtp_user

    if not smtp_host or not smtp_user or not smtp_pass:
        print('ERRO: SMTP não configurado no .env')
        sys.exit(1)

    msg = MIMEMultipart('mixed')
    msg['Subject'] = f"[TESTE] Movimentações Padioleiro HAC — {now.strftime('%d/%m/%Y %H:%M')}"
    msg['From']    = f"Painel HAC <{smtp_from}>"
    msg['To']      = destinatario
    msg.attach(MIMEText(html, 'html', 'utf-8'))

    nome_excel = f"Padioleiro_HAC_{now.strftime('%d-%m-%Y_%H%M')}.xlsx"
    with open(excel_path, 'rb') as f:
        part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename="{nome_excel}"')
    msg.attach(part)

    print(f'Conectando em {smtp_host}:{smtp_port}...')
    server = smtplib.SMTP(smtp_host, smtp_port, timeout=30)
    server.ehlo(); server.starttls(); server.login(smtp_user, smtp_pass)
    server.sendmail(smtp_from, [destinatario], msg.as_bytes())
    server.quit()
    print(f'Email + Excel enviados para {destinatario}')


# ── MAIN ──────────────────────────────────────────────────────

if __name__ == '__main__':
    now = datetime.now()
    print(f'Buscando dados de {date.today().strftime("%d/%m/%Y")}...')

    try:
        resumo, por_padioleiro, por_setor, por_hora, tendencia_7d, todos_hoje, cancelados = buscar_dados()
    except Exception as e:
        print(f'ERRO ao consultar banco: {e}')
        sys.exit(1)

    print(f'  Total de chamados hoje : {resumo.get("total", 0)}')
    print(f'  Concluidos             : {resumo.get("concluidos", 0)}')
    print(f'  Cancelados             : {resumo.get("cancelados", 0)}')
    print(f'  Em aberto              : {resumo.get("em_aberto", 0)}')
    print(f'  Padioleiros com mov.   : {len(por_padioleiro)}')
    print(f'  Setores com mov.       : {len(por_setor)}')
    print(f'  Tendência (dias)       : {len(tendencia_7d)}')
    print()

    print('Gerando Excel...')
    excel_path = None
    try:
        excel_path = gerar_excel(resumo, por_padioleiro, por_setor, por_hora, tendencia_7d, todos_hoje, cancelados, now)
        print(f'  Excel gerado: {excel_path}')
    except Exception as e:
        print(f'ERRO ao gerar Excel: {e}')
        sys.exit(1)

    html = gerar_html(resumo, por_padioleiro, por_setor, cancelados, now)

    try:
        enviar(DESTINATARIO, html, excel_path, now)
    except Exception as e:
        print(f'ERRO ao enviar email: {e}')
        sys.exit(1)
    finally:
        if excel_path and os.path.exists(excel_path):
            os.unlink(excel_path)

"""
========================================
WORKER - NOTIFICADOR DE OCUPAÇÃO HOSPITALAR
========================================

Envia relatório de ocupação hospitalar por email em intervalos configuráveis.

Configuração via .env:
  NOTIF_OCUPACAO_EMAILS      - destinatários separados por vírgula
  NOTIF_OCUPACAO_INTERVALO_H - intervalo em horas (padrão: 6)
  NOTIF_OCUPACAO_HORARIOS    - horários fixos separados por vírgula, ex: 06:00,12:00,18:00,00:00
                               (se definido, substitui o intervalo)

O email contém:
  - Corpo HTML com resumo geral da ocupação
  - Anexo Excel com 3 abas: Resumo, Por Setor, Pacientes Internados
"""

import os
import sys
import time
import logging
import smtplib
import tempfile
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

import psycopg2
from psycopg2.extras import RealDictCursor
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ========================================
# CONFIGURAÇÃO
# ========================================

_BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
load_dotenv(os.path.join(_BASE_DIR, '.env'))
os.makedirs(os.path.join(_BASE_DIR, 'logs'), exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] %(levelname)s: %(message)s',
    handlers=[
        logging.FileHandler(
            os.path.join(_BASE_DIR, 'logs', 'notificador_ocupacao.log'),
            encoding='utf-8'
        ),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


def _cfg():
    emails_raw = os.getenv('NOTIF_OCUPACAO_EMAILS', '')
    horarios_raw = os.getenv('NOTIF_OCUPACAO_HORARIOS', '')
    return {
        'destinatarios': [e.strip() for e in emails_raw.split(',') if e.strip()],
        'horarios':      [h.strip() for h in horarios_raw.split(',') if h.strip()],
        'intervalo_h':   float(os.getenv('NOTIF_OCUPACAO_INTERVALO_H', '6')),
    }


# ========================================
# BANCO DE DADOS
# ========================================

def _get_conn():
    return psycopg2.connect(
        host=os.getenv('DB_HOST', 'localhost'),
        dbname=os.getenv('DB_NAME', 'postgres'),
        user=os.getenv('DB_USER', 'postgres'),
        password=os.getenv('DB_PASSWORD', ''),
        port=int(os.getenv('DB_PORT', '5432')),
        connect_timeout=10
    )


def _query(conn, sql):
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute(sql)
        rows = cur.fetchall()
        cols = [desc.name for desc in cur.description] if cur.description else []
    return cols, [dict(r) for r in rows]


def buscar_dados():
    conn = _get_conn()
    try:
        _, dash_rows = _query(conn, "SELECT * FROM vw_ocupacao_dashboard")
        dashboard = dash_rows[0] if dash_rows else {}

        cols_setor, setores = _query(conn, "SELECT * FROM vw_ocupacao_por_setor")
        cols_pac,   pacientes = _query(conn, "SELECT * FROM vw_pacientes_internados ORDER BY 1, 2")

        return dashboard, cols_setor, setores, cols_pac, pacientes
    finally:
        conn.close()


# ========================================
# UTILITÁRIOS
# ========================================

# Colunas que têm o nome do setor — o Apache Hop grava a expressão como nome de coluna
_CANDIDATOS_SETOR = [
    'obter_nome_setor(a.cd_setor_atendimento)',
    'nm_setor', 'nome_setor', 'ds_setor', 'descricao_setor', 'setor',
]

# Colunas de nome de pessoa (LGPD)
_COLUNAS_NOME = {'nm_pessoa_fisica', 'nm_guerra'}

# Mapeamento de nomes de colunas técnicos para labels legíveis no Excel
_COL_LABELS = {
    'obter_nome_setor(a.cd_setor_atendimento)': 'Setor',
    'cd_setor_atendimento': 'Cód. Setor',
    'nm_pessoa_fisica':     'Paciente',
    'nm_guerra':            'Médico',
    'nr_atendimento':       'Atendimento',
    'dt_entrada_unidade':   'Entrada',
    'qt_dia_permanencia':   'Dias',
    'ds_convenio':          'Convênio',
    'ds_clinica':           'Clínica',
    'ie_status_unidade':    'Status',
    'ie_sexo':              'Sexo',
}


def _label_coluna(col):
    return _COL_LABELS.get(col.lower(), col.replace('_', ' ').title())


def _nome_setor(row):
    """Extrai o nome do setor buscando pelas colunas candidatas em ordem de prioridade."""
    for c in _CANDIDATOS_SETOR:
        val = row.get(c)
        if val and str(val).strip():
            return str(val).strip()
    return '-'


def _anonimizar(nome):
    """
    LGPD: Iniciais de todos os nomes exceto o último, que fica completo.
    'Lucas Fernandes de Oliveira' → 'L F D Oliveira'
    """
    if not nome:
        return nome
    partes = str(nome).strip().split()
    if len(partes) == 1:
        return partes[0][0].upper() + '.'
    iniciais = [p[0].upper() for p in partes[:-1]]
    return ' '.join(iniciais) + ' ' + partes[-1].capitalize()


# ========================================
# CORES
# ========================================

_COR_HEADER  = "9B1C24"   # vermelho escuro
_COR_TITULO  = "DC3545"   # vermelho padrão HAC
_COR_CRITICO = "C00000"   # vermelho crítico
_COR_LIVRE   = "375623"   # verde
_COR_ALERTA  = "FF8C00"   # laranja
_COR_BRANCO  = "FFFFFF"
_COR_ZEBRA   = "FDECEA"   # rosa claro


def _cor_taxa(valor):
    try:
        v = float(valor)
        if v >= 90:
            return "#C00000", "CRÍTICA"
        if v >= 75:
            return "#FF8C00", "ALERTA"
        return "#375623", "NORMAL"
    except (TypeError, ValueError):
        return "#555555", "-"


# ========================================
# EXCEL
# ========================================

def _hdr(cell, cor=_COR_HEADER):
    cell.font = Font(bold=True, color=_COR_BRANCO, size=11)
    cell.fill = PatternFill("solid", fgColor=cor)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _borda(cell)


def _borda(cell):
    s = Side(style="thin", color="AAAAAA")
    cell.border = Border(left=s, right=s, top=s, bottom=s)


def _autowidth(ws, min_w=10, max_w=50):
    for col in ws.columns:
        letra = get_column_letter(col[0].column)
        w = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[letra].width = min(max(w + 2, min_w), max_w)


def _aba_resumo(wb, dashboard, now):
    ws = wb.create_sheet("Resumo Geral")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:B1")
    t = ws["A1"]
    t.value = f"OCUPAÇÃO HOSPITALAR — HAC — {now.strftime('%d/%m/%Y %H:%M')}"
    t.font = Font(bold=True, color=_COR_BRANCO, size=14)
    t.fill = PatternFill("solid", fgColor=_COR_TITULO)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.append([])

    metricas = [
        ("Indicador",                  "Valor"),
        ("Total de Leitos",            dashboard.get('total_leitos', '-')),
        ("Leitos Ocupados",            dashboard.get('leitos_ocupados', '-')),
        ("Leitos Livres",              dashboard.get('leitos_livres', '-')),
        ("Em Higienização",            dashboard.get('leitos_higienizacao', '-')),
        ("Interditados",               dashboard.get('leitos_interditados', '-')),
        ("Taxa de Ocupação (%)",       dashboard.get('taxa_ocupacao_geral', '-')),
        ("Taxa de Disponibilidade (%)", dashboard.get('taxa_disponibilidade', '-') or '-'),
        ("Total de Setores",           dashboard.get('total_setores', '-')),
        ("Média de Permanência (dias)", dashboard.get('media_permanencia_geral', '-')),
        ("Última Atualização",         str(dashboard.get('ultima_atualizacao', '-'))),
    ]

    for i, (indicador, valor) in enumerate(metricas):
        r = i + 3
        c1 = ws.cell(row=r, column=1, value=indicador)
        c2 = ws.cell(row=r, column=2, value=valor)
        if i == 0:
            _hdr(c1); _hdr(c2)
        else:
            c1.font = Font(bold=True)
            c1.fill = PatternFill("solid", fgColor="F5D0D3")
            _borda(c1); _borda(c2)
            if indicador == "Taxa de Ocupação (%)":
                try:
                    v = float(valor)
                    cor = _COR_CRITICO if v >= 90 else (_COR_ALERTA if v >= 75 else _COR_LIVRE)
                    c2.font = Font(bold=True, color=_COR_BRANCO)
                    c2.fill = PatternFill("solid", fgColor=cor)
                except (TypeError, ValueError):
                    pass

    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 22


def _aba_por_setor(wb, cols, setores):
    ws = wb.create_sheet("Por Setor")
    ws.sheet_view.showGridLines = False
    if not cols:
        ws["A1"] = "Sem dados"; return

    for j, col in enumerate(cols, 1):
        _hdr(ws.cell(row=1, column=j, value=_label_coluna(col)))
    ws.row_dimensions[1].height = 22

    taxa_idx = next((i for i, c in enumerate(cols) if 'taxa' in c.lower() and 'ocup' in c.lower()), None)

    for i, row in enumerate(setores, 2):
        zebra = i % 2 == 0
        for j, col in enumerate(cols, 1):
            val = row.get(col)
            # Usar nome do setor quando for a coluna de setor
            if col.lower() in _CANDIDATOS_SETOR:
                val = _nome_setor(row)
            cell = ws.cell(row=i, column=j, value=val)
            _borda(cell)
            cell.alignment = Alignment(vertical="center")
            if zebra:
                cell.fill = PatternFill("solid", fgColor=_COR_ZEBRA)
            if taxa_idx is not None and j == taxa_idx + 1:
                try:
                    v = float(row.get(col, 0) or 0)
                    cor = _COR_CRITICO if v >= 90 else (_COR_ALERTA if v >= 75 else None)
                    if cor:
                        cell.font = Font(bold=True, color=_COR_BRANCO)
                        cell.fill = PatternFill("solid", fgColor=cor)
                except (TypeError, ValueError):
                    pass

    _autowidth(ws)


def _aba_pacientes(wb, cols, pacientes):
    ws = wb.create_sheet("Pacientes Internados")
    ws.sheet_view.showGridLines = False
    if not cols:
        ws["A1"] = "Sem dados"; return

    for j, col in enumerate(cols, 1):
        _hdr(ws.cell(row=1, column=j, value=_label_coluna(col)))
    ws.row_dimensions[1].height = 22

    for i, row in enumerate(pacientes, 2):
        zebra = i % 2 == 0
        for j, col in enumerate(cols, 1):
            val = row.get(col)
            if isinstance(val, datetime):
                val = val.strftime('%d/%m/%Y %H:%M')
            # Nome do setor
            if col.lower() in _CANDIDATOS_SETOR:
                val = _nome_setor(row)
            # Anonimização LGPD
            elif col.lower() in _COLUNAS_NOME and val:
                val = _anonimizar(val)
            cell = ws.cell(row=i, column=j, value=val)
            _borda(cell)
            cell.alignment = Alignment(vertical="center")
            if zebra:
                cell.fill = PatternFill("solid", fgColor=_COR_ZEBRA)

    _autowidth(ws)


def gerar_excel(dashboard, cols_setor, setores, cols_pac, pacientes, now):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _aba_resumo(wb, dashboard, now)
    _aba_por_setor(wb, cols_setor, setores)
    _aba_pacientes(wb, cols_pac, pacientes)

    tmp = tempfile.NamedTemporaryFile(
        delete=False, suffix='.xlsx',
        prefix=f"ocupacao_hac_{now.strftime('%Y%m%d_%H%M')}_"
    )
    wb.save(tmp.name)
    tmp.close()
    return tmp.name


# ========================================
# EMAIL HTML
# ========================================

def gerar_corpo_html(dashboard, setores, now):
    taxa      = dashboard.get('taxa_ocupacao_geral', 0) or 0
    cor_taxa, status_taxa = _cor_taxa(taxa)
    ocupados  = dashboard.get('leitos_ocupados', 0) or 0
    total     = dashboard.get('total_leitos', 0) or 0
    livres    = dashboard.get('leitos_livres', 0) or 0
    higieniz  = dashboard.get('leitos_higienizacao', 0) or 0
    interdit  = dashboard.get('leitos_interditados', 0) or 0
    setores_n = dashboard.get('total_setores', 0) or 0
    perm      = dashboard.get('media_permanencia_geral', 0) or 0

    linhas = ""
    for i, s in enumerate(setores):
        bg      = "#FDECEA" if i % 2 == 0 else "#FFFFFF"
        nome    = _nome_setor(s)
        t_leitos = s.get('total_leitos') or s.get('leitos') or '-'
        t_ocup   = s.get('leitos_ocupados') or s.get('ocupados') or '-'
        t_livre  = s.get('leitos_livres') or s.get('livres') or '-'
        t_taxa   = s.get('taxa_ocupacao') or s.get('taxa_ocupacao_geral') or '-'
        cor_s, _ = _cor_taxa(t_taxa)
        taxa_fmt = f"{t_taxa}%" if t_taxa != '-' else '-'
        linhas += f"""
        <tr style="background:{bg}">
          <td style="padding:6px 10px;border:1px solid #ddd">{nome}</td>
          <td style="padding:6px 10px;border:1px solid #ddd;text-align:center">{t_leitos}</td>
          <td style="padding:6px 10px;border:1px solid #ddd;text-align:center">{t_ocup}</td>
          <td style="padding:6px 10px;border:1px solid #ddd;text-align:center">{t_livre}</td>
          <td style="padding:6px 10px;border:1px solid #ddd;text-align:center;color:{cor_s};font-weight:bold">{taxa_fmt}</td>
        </tr>"""

    tabela = f"""
    <table style="width:100%;border-collapse:collapse;margin-top:8px;font-size:13px">
      <thead>
        <tr style="background:#dc3545;color:#fff">
          <th style="padding:8px 10px;border:1px solid #ddd;text-align:left">Setor</th>
          <th style="padding:8px 10px;border:1px solid #ddd">Total</th>
          <th style="padding:8px 10px;border:1px solid #ddd">Ocupados</th>
          <th style="padding:8px 10px;border:1px solid #ddd">Livres</th>
          <th style="padding:8px 10px;border:1px solid #ddd">Taxa</th>
        </tr>
      </thead>
      <tbody>{linhas}</tbody>
    </table>""" if setores else "<p style='color:#888'>Sem dados de setores.</p>"

    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"></head>
<body style="font-family:Arial,sans-serif;margin:0;padding:0;background:#f4f4f4">
<div style="max-width:700px;margin:20px auto;background:#fff;border-radius:8px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,0.15)">

  <div style="background:#dc3545;padding:24px 30px;color:#fff">
    <h1 style="margin:0;font-size:20px">Relatório de Ocupação Hospitalar</h1>
    <p style="margin:4px 0 0;font-size:13px;opacity:0.85">Hospital Anchieta Ceilândia — {now.strftime('%d/%m/%Y às %H:%M')}</p>
  </div>

  <div style="padding:24px 30px">
    <div style="display:flex;gap:12px;flex-wrap:wrap;margin-bottom:20px">

      <div style="flex:1;min-width:120px;background:#fff5f5;border-radius:8px;padding:16px;text-align:center;border-top:4px solid {cor_taxa}">
        <div style="font-size:32px;font-weight:bold;color:{cor_taxa}">{taxa}%</div>
        <div style="font-size:12px;color:#555;margin-top:4px">Taxa de Ocupação</div>
        <div style="font-size:11px;font-weight:bold;color:{cor_taxa}">{status_taxa}</div>
      </div>

      <div style="flex:1;min-width:120px;background:#fff0f0;border-radius:8px;padding:16px;text-align:center;border-top:4px solid #C00000">
        <div style="font-size:32px;font-weight:bold;color:#C00000">{ocupados}</div>
        <div style="font-size:12px;color:#555;margin-top:4px">Leitos Ocupados</div>
        <div style="font-size:11px;color:#888">de {total} no total</div>
      </div>

      <div style="flex:1;min-width:120px;background:#f0fff4;border-radius:8px;padding:16px;text-align:center;border-top:4px solid #375623">
        <div style="font-size:32px;font-weight:bold;color:#375623">{livres}</div>
        <div style="font-size:12px;color:#555;margin-top:4px">Leitos Livres</div>
      </div>

      <div style="flex:1;min-width:120px;background:#fffaf0;border-radius:8px;padding:16px;text-align:center;border-top:4px solid #FF8C00">
        <div style="font-size:24px;font-weight:bold;color:#FF8C00">{higieniz} / {interdit}</div>
        <div style="font-size:12px;color:#555;margin-top:4px">Higienização / Interditados</div>
      </div>

    </div>

    <div style="background:#f9f9f9;border-radius:6px;padding:12px 16px;margin-bottom:20px;font-size:13px;color:#444">
      <strong>{setores_n}</strong> setores monitorados &nbsp;|&nbsp;
      Média de permanência: <strong>{perm} dias</strong>
    </div>

    <h2 style="font-size:15px;color:#dc3545;margin:0 0 8px">Ocupação por Setor</h2>
    {tabela}

    <p style="font-size:12px;color:#888;margin-top:20px;border-top:1px solid #eee;padding-top:12px">
      O arquivo Excel em anexo contém a lista completa de pacientes internados.<br>
      Nomes anonimizados conforme LGPD. Gerado automaticamente pelo Sistema de Painéis — HAC.
    </p>
  </div>

</div>
</body></html>"""


# ========================================
# ENVIO DE EMAIL
# ========================================

def enviar_email(destinatarios, corpo_html, caminho_excel, now):
    smtp_host = os.getenv('SMTP_HOST', '')
    smtp_port = int(os.getenv('SMTP_PORT', '587'))
    smtp_user = os.getenv('SMTP_USER', '')
    smtp_pass = os.getenv('SMTP_PASS', '')
    smtp_from = os.getenv('SMTP_FROM', '') or smtp_user

    if not smtp_host or not smtp_user or not smtp_pass:
        logger.error("SMTP não configurado no .env")
        return False

    nome_arquivo = f"Ocupacao_HAC_{now.strftime('%d-%m-%Y_%H%M')}.xlsx"
    msg = MIMEMultipart('mixed')
    msg['Subject'] = f"Ocupação Hospitalar HAC — {now.strftime('%d/%m/%Y %H:%M')}"
    msg['From']    = f"Painel HAC <{smtp_from}>"
    msg['To']      = ', '.join(destinatarios)
    msg.attach(MIMEText(corpo_html, 'html', 'utf-8'))

    with open(caminho_excel, 'rb') as f:
        part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename="{nome_arquivo}"')
    msg.attach(part)

    try:
        server = smtplib.SMTP(smtp_host, smtp_port, timeout=30)
        server.ehlo()
        server.starttls()
        server.login(smtp_user, smtp_pass)
        server.sendmail(smtp_from, destinatarios, msg.as_bytes())
        server.quit()
        logger.info("Email enviado para: %s", ', '.join(destinatarios))
        return True
    except Exception as e:
        logger.error("Erro ao enviar email: %s", e)
        return False


# ========================================
# CICLO PRINCIPAL
# ========================================

def executar_envio():
    cfg = _cfg()
    if not cfg['destinatarios']:
        logger.warning("Nenhum destinatário em NOTIF_OCUPACAO_EMAILS. Pulando.")
        return

    now = datetime.now()
    logger.info("Iniciando coleta de dados — %s", now.strftime('%d/%m/%Y %H:%M'))

    try:
        dashboard, cols_setor, setores, cols_pac, pacientes = buscar_dados()
    except Exception as e:
        logger.error("Erro ao buscar dados do banco: %s", e)
        return

    caminho = None
    try:
        caminho = gerar_excel(dashboard, cols_setor, setores, cols_pac, pacientes, now)
        html    = gerar_corpo_html(dashboard, setores, now)
        enviar_email(cfg['destinatarios'], html, caminho, now)
    except Exception as e:
        logger.error("Erro ao gerar/enviar relatório: %s", e)
    finally:
        if caminho and os.path.exists(caminho):
            os.unlink(caminho)


def _proximo_horario(horarios):
    now = datetime.now()
    candidatos = []
    for h in horarios:
        try:
            hora, minuto = map(int, h.split(':'))
        except ValueError:
            continue
        c = now.replace(hour=hora, minute=minuto, second=0, microsecond=0)
        if c <= now:
            c += timedelta(days=1)
        candidatos.append(c)
    return min(candidatos) if candidatos else None


def main():
    logger.info("=" * 50)
    logger.info("Notificador de Ocupação Hospitalar — Iniciando")
    logger.info("=" * 50)

    cfg = _cfg()
    logger.info("Destinatários : %s", cfg['destinatarios'] or '(nenhum)')
    logger.info("Horários fixos: %s", cfg['horarios'] or '(não configurado)')
    logger.info("Intervalo     : %sh", cfg['intervalo_h'])

    if not cfg['destinatarios']:
        logger.error("Configure NOTIF_OCUPACAO_EMAILS no .env e reinicie.")
        sys.exit(1)

    executar_envio()

    while True:
        cfg = _cfg()
        proximo = _proximo_horario(cfg['horarios']) if cfg['horarios'] else \
                  datetime.now() + timedelta(hours=cfg['intervalo_h'])
        if proximo is None:
            proximo = datetime.now() + timedelta(hours=6)

        logger.info("Próximo envio: %s", proximo.strftime('%d/%m/%Y %H:%M'))
        while datetime.now() < proximo:
            time.sleep(30)
        executar_envio()


# ========================================
# INTEGRAÇÃO FLASK/GUNICORN (thread daemon)
# ========================================

_background_started = False


def start_in_background():
    """
    Inicia o notificador como thread daemon junto com o Flask/Gunicorn.
    OFF SWITCH: NOTIF_OCUPACAO_AUTO=false no .env
    """
    global _background_started
    if _background_started:
        return

    if os.getenv('NOTIF_OCUPACAO_AUTO', 'true').lower() != 'true':
        logger.info('[notificador_ocupacao] Auto-start desativado (NOTIF_OCUPACAO_AUTO=false)')
        return

    try:
        from werkzeug.serving import is_running_from_reloader
        if is_running_from_reloader() and os.environ.get('WERKZEUG_RUN_MAIN') != 'true':
            return
    except ImportError:
        pass

    _background_started = True
    import threading

    def _run():
        try:
            cfg = _cfg()
            logger.info(
                '[notificador_ocupacao] Thread daemon iniciada — destinatários: %s | horários: %s | intervalo: %sh',
                cfg['destinatarios'], cfg['horarios'], cfg['intervalo_h']
            )
            time.sleep(60)  # aguarda Flask terminar de subir
            executar_envio()

            while True:
                cfg = _cfg()
                proximo = _proximo_horario(cfg['horarios']) if cfg['horarios'] else \
                          datetime.now() + timedelta(hours=cfg['intervalo_h'])
                if proximo is None:
                    proximo = datetime.now() + timedelta(hours=6)
                logger.info('[notificador_ocupacao] Próximo envio: %s', proximo.strftime('%d/%m/%Y %H:%M'))
                while datetime.now() < proximo:
                    time.sleep(30)
                executar_envio()

        except Exception as e:
            logger.error('[notificador_ocupacao] Erro fatal na thread daemon: %s', e, exc_info=True)

    t = threading.Thread(target=_run, name='notificador_ocupacao', daemon=True)
    t.start()
    logger.info('[notificador_ocupacao] Thread daemon registrada')


if __name__ == '__main__':
    main()

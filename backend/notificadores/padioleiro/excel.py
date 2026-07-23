# -*- coding: utf-8 -*-
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference, PieChart
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import ColorScaleRule
from .utils import _data_pt

_HAC      = "9B1C24"
_VERDE    = "28A745"
_VERMELHO = "DC3545"
_LARANJA  = "E67E00"
_AZUL     = "17A2B8"
_BRANCO   = "FFFFFF"
_ZEBRA    = "FEF0F0"
_FUNDO_H  = "F5D0D3"


def _hdr(cell, cor=_HAC):
    cell.font = Font(bold=True, color=_BRANCO, size=11)
    cell.fill = PatternFill("solid", fgColor=cor)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _borda(cell)


def _borda(cell):
    s = Side(style="thin", color="CCCCCC")
    cell.border = Border(left=s, right=s, top=s, bottom=s)


def _autowidth(ws, min_w=10, max_w=45):
    for col in ws.columns:
        letra = get_column_letter(col[0].column)
        w = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[letra].width = min(max(w + 3, min_w), max_w)


def _titulo(ws, texto, ncols, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=texto)
    c.font = Font(bold=True, color=_BRANCO, size=13)
    c.fill = PatternFill("solid", fgColor=_HAC)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28


def _serie_cor(serie, cor):
    serie.graphicalProperties.solidFill = cor


def _cor_tempo(ws, col_letra, row_ini, row_fim):
    rng = f"{col_letra}{row_ini}:{col_letra}{row_fim}"
    ws.conditional_formatting.add(rng, ColorScaleRule(
        start_type='num', start_value=0,  start_color="63BE7B",
        mid_type='num',   mid_value=20,   mid_color="FFEB84",
        end_type='num',   end_value=60,   end_color="F8696B",
    ))


def _aba_resumo(wb, resumo, now, nome_periodo):
    ws = wb.create_sheet("Resumo do Dia")
    ws.sheet_view.showGridLines = False
    _titulo(ws, f"MOVIMENTACOES DO PADIOLEIRO — HAC — {_data_pt(now)} — {nome_periodo}", 5)

    total      = int(resumo.get('total') or 0)
    concluidos = int(resumo.get('concluidos') or 0)
    cancelados = int(resumo.get('cancelados') or 0)
    em_aberto  = int(resumo.get('em_aberto') or 0)
    urgentes   = int(resumo.get('urgentes') or 0)
    taxa       = round(concluidos / total * 100) if total else 0

    kpis = [
        ("Indicador", "Valor", None),
        ("Total de Chamados",            total,                                  _HAC),
        ("Concluidos",                   concluidos,                             _VERDE),
        ("Cancelados",                   cancelados,                             _VERMELHO),
        ("Em Aberto",                    em_aberto,                              _AZUL),
        ("Urgentes",                     urgentes,                               _LARANJA),
        ("Taxa de Conclusao (%)",        taxa,                                   _HAC),
        ("Tempo Medio ate Aceite (min)", resumo.get('media_aceite_min') or '--', None),
        ("Tempo Medio Total (min)",      resumo.get('media_total_min') or '--',  None),
    ]
    for i, (ind, val, cor) in enumerate(kpis, 3):
        c1 = ws.cell(row=i, column=1, value=ind)
        c2 = ws.cell(row=i, column=2, value=val)
        if i == 3:
            _hdr(c1); _hdr(c2)
        else:
            c1.font = Font(bold=True)
            c1.fill = PatternFill("solid", fgColor=_FUNDO_H)
            _borda(c1); _borda(c2)
            if cor:
                c2.font = Font(bold=True, color=_BRANCO)
                c2.fill = PatternFill("solid", fgColor=cor)

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18

    pie_items = [("Status", "Qtd"), ("Concluidos", concluidos), ("Cancelados", cancelados), ("Em Aberto", em_aberto)]
    for i, (s, v) in enumerate(pie_items, 3):
        ws.cell(row=i, column=4, value=s).font = Font(bold=(i == 3))
        ws.cell(row=i, column=5, value=v)
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 10

    if total > 0:
        pie = PieChart()
        pie.title = "Distribuicao por Status"
        pie.style = 10; pie.height = 12; pie.width = 16
        pie.add_data(Reference(ws, min_col=5, max_col=5, min_row=3, max_row=6), titles_from_data=True)
        pie.set_categories(Reference(ws, min_col=4, min_row=4, max_row=6))
        for idx, cor in enumerate([_VERDE, _VERMELHO, _AZUL]):
            pt = DataPoint(idx=idx)
            pt.graphicalProperties.solidFill = cor
            pie.series[0].dPt.append(pt)
        ws.add_chart(pie, "D8")


def _aba_por_padioleiro(wb, por_padioleiro):
    ws = wb.create_sheet("Por Padioleiro")
    ws.sheet_view.showGridLines = False

    hv = ["Padioleiro", "Total", "Concluidos", "Cancelados", "Urgentes"]
    _titulo(ws, "VOLUMES POR PADIOLEIRO", len(hv), 1)
    for j, h in enumerate(hv, 1):
        _hdr(ws.cell(row=2, column=j, value=h))

    for i, p in enumerate(por_padioleiro, 3):
        zebra = i % 2 == 0
        for j, (key, cor_txt) in enumerate([
            ('padioleiro', None), ('total', None), ('concluidos', _VERDE),
            ('cancelados', _VERMELHO), ('urgentes', _LARANJA)
        ], 1):
            cell = ws.cell(row=i, column=j, value=p.get(key))
            _borda(cell)
            if zebra: cell.fill = PatternFill("solid", fgColor=_ZEBRA)
            if j == 1: cell.font = Font(bold=True)
            elif cor_txt: cell.font = Font(bold=True, color=cor_txt)

    last_vol = 2 + len(por_padioleiro)

    if por_padioleiro:
        c = BarChart()
        c.type = "bar"; c.grouping = "clustered"
        c.title = "Movimentos por Padioleiro"
        c.style = 10; c.height = 12; c.width = 22
        c.x_axis.title = "Quantidade"
        c.add_data(Reference(ws, min_col=2, max_col=5, min_row=2, max_row=last_vol), titles_from_data=True)
        c.set_categories(Reference(ws, min_col=1, min_row=3, max_row=last_vol))
        for idx, cor in enumerate([_HAC, _VERDE, _VERMELHO, _LARANJA]):
            if idx < len(c.series): _serie_cor(c.series[idx], cor)
        ws.add_chart(c, f"A{last_vol + 3}")
        prox = last_vol + 30
    else:
        prox = last_vol + 3

    ht = ["Padioleiro", "T.Aceite(min)", "T.Desloc.(min)", "T.Transp.(min)", "T.Total(min)"]
    _titulo(ws, "TEMPOS MEDIOS POR PADIOLEIRO", len(ht), prox)
    for j, h in enumerate(ht, 1):
        _hdr(ws.cell(row=prox + 1, column=j, value=h))

    for i, p in enumerate(por_padioleiro, prox + 2):
        zebra = i % 2 == 0
        for j, key in enumerate(['padioleiro','media_aceite_min','media_deslocamento_min','media_transporte_min','media_total_min'], 1):
            v = p.get(key)
            cell = ws.cell(row=i, column=j, value=float(v) if v is not None and j > 1 else v)
            _borda(cell)
            if zebra: cell.fill = PatternFill("solid", fgColor=_ZEBRA)
            if j == 1: cell.font = Font(bold=True)

    last_tem = prox + 1 + len(por_padioleiro)
    if por_padioleiro:
        for col_l in ['B', 'C', 'D', 'E']:
            _cor_tempo(ws, col_l, prox + 2, last_tem)
        c2 = BarChart()
        c2.type = "bar"; c2.grouping = "clustered"
        c2.title = "Tempos Medios por Padioleiro (min)"
        c2.style = 10; c2.height = 12; c2.width = 22
        c2.x_axis.title = "Minutos"
        c2.add_data(Reference(ws, min_col=2, max_col=5, min_row=prox+1, max_row=last_tem), titles_from_data=True)
        c2.set_categories(Reference(ws, min_col=1, min_row=prox+2, max_row=last_tem))
        for idx, cor in enumerate([_AZUL, _HAC, _VERDE, _LARANJA]):
            if idx < len(c2.series): _serie_cor(c2.series[idx], cor)
        ws.add_chart(c2, f"A{last_tem + 3}")

    _autowidth(ws)


def _aba_por_setor(wb, por_setor):
    ws = wb.create_sheet("Por Setor")
    ws.sheet_view.showGridLines = False
    headers = ["Setor", "Total", "Concluidos", "Cancelados", "Urgentes"]
    _titulo(ws, "CHAMADOS POR SETOR DE ORIGEM", len(headers))
    for j, h in enumerate(headers, 1):
        _hdr(ws.cell(row=2, column=j, value=h))

    for i, s in enumerate(por_setor, 3):
        zebra = i % 2 == 0
        for j, (key, cor_txt) in enumerate([
            ('setor',None),('total',None),('concluidos',_VERDE),('cancelados',_VERMELHO),('urgentes',_LARANJA)
        ], 1):
            cell = ws.cell(row=i, column=j, value=s.get(key))
            _borda(cell)
            if zebra: cell.fill = PatternFill("solid", fgColor=_ZEBRA)
            if j == 1: cell.font = Font(bold=True)
            elif cor_txt: cell.font = Font(bold=True, color=cor_txt)

    last_row = 2 + len(por_setor)
    if por_setor:
        c = BarChart()
        c.type = "bar"; c.grouping = "clustered"
        c.title = "Chamados por Setor"
        c.style = 10; c.height = max(12, len(por_setor) * 0.9); c.width = 22
        c.x_axis.title = "Quantidade"
        c.add_data(Reference(ws, min_col=2, max_col=5, min_row=2, max_row=last_row), titles_from_data=True)
        c.set_categories(Reference(ws, min_col=1, min_row=3, max_row=last_row))
        for idx, cor in enumerate([_HAC, _VERDE, _VERMELHO, _LARANJA]):
            if idx < len(c.series): _serie_cor(c.series[idx], cor)
        ws.add_chart(c, f"A{last_row + 3}")
    _autowidth(ws)


def _aba_por_hora(wb, por_hora, horas_periodo, nome_periodo):
    ws = wb.create_sheet("Distribuicao por Hora")
    ws.sheet_view.showGridLines = False
    headers = ["Hora", "Total", "Concluidos", "Urgentes"]
    _titulo(ws, f"CHAMADOS POR HORA — {nome_periodo}", len(headers))
    for j, h in enumerate(headers, 1):
        _hdr(ws.cell(row=2, column=j, value=h))

    hora_dict = {r['hora']: r for r in por_hora}
    for idx, hora in enumerate(horas_periodo):
        r = hora_dict.get(hora, {})
        row_n = idx + 3
        zebra = idx % 2 == 0
        for j, v in enumerate([f"{hora:02d}h", int(r.get('total',0)), int(r.get('concluidos',0)), int(r.get('urgentes',0))], 1):
            cell = ws.cell(row=row_n, column=j, value=v)
            _borda(cell)
            if zebra: cell.fill = PatternFill("solid", fgColor=_ZEBRA)

    max_row_dados = 2 + len(horas_periodo)
    c = BarChart()
    c.type = "col"; c.grouping = "clustered"
    c.title = f"Distribuicao de Chamados — {nome_periodo}"
    c.style = 10; c.height = 14; c.width = 26
    c.y_axis.title = "Chamados"; c.x_axis.title = "Hora"
    c.add_data(Reference(ws, min_col=2, max_col=4, min_row=2, max_row=max_row_dados), titles_from_data=True)
    c.set_categories(Reference(ws, min_col=1, min_row=3, max_row=max_row_dados))
    for idx, cor in enumerate([_HAC, _VERDE, _LARANJA]):
        if idx < len(c.series): _serie_cor(c.series[idx], cor)
    ws.add_chart(c, "F2")
    _autowidth(ws)


def _aba_tendencia(wb, tendencia_7d):
    ws = wb.create_sheet("Tendencia 7 Dias")
    ws.sheet_view.showGridLines = False
    headers = ["Data", "Total", "Concluidos", "Cancelados", "T.Medio(min)"]
    _titulo(ws, "TENDENCIA — ULTIMOS 7 DIAS", len(headers))
    for j, h in enumerate(headers, 1):
        _hdr(ws.cell(row=2, column=j, value=h))

    for i, d in enumerate(tendencia_7d, 3):
        zebra = i % 2 == 0
        dt = d.get('data')
        data_fmt = dt.strftime('%d/%m') if hasattr(dt, 'strftime') else str(dt or '')
        for j, v in enumerate([data_fmt, d.get('total'), d.get('concluidos'), d.get('cancelados'), d.get('media_total_min')], 1):
            cell = ws.cell(row=i, column=j, value=float(v) if v is not None and j == 5 else v)
            _borda(cell)
            if zebra: cell.fill = PatternFill("solid", fgColor=_ZEBRA)

    last_row = 2 + len(tendencia_7d)
    if tendencia_7d:
        c = LineChart()
        c.title = "Tendencia de Chamados — Ultimos 7 Dias"
        c.style = 10; c.height = 14; c.width = 22
        c.y_axis.title = "Chamados"; c.x_axis.title = "Data"
        c.add_data(Reference(ws, min_col=2, max_col=4, min_row=2, max_row=last_row), titles_from_data=True)
        c.set_categories(Reference(ws, min_col=1, min_row=3, max_row=last_row))
        for idx, cor in enumerate([_HAC, _VERDE, _VERMELHO]):
            if idx < len(c.series):
                c.series[idx].graphicalProperties.line.solidFill = cor
        ws.add_chart(c, f"A{last_row + 3}")
    _autowidth(ws)


def _aba_todos_chamados(wb, todos_hoje, now, nome_periodo):
    ws = wb.create_sheet("Todos os Chamados")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    headers = [
        "#", "Tipo", "Paciente", "Leito", "Setor Origem", "Destino",
        "Prioridade", "Status", "Solicitante", "Padioleiro",
        "Criado", "Aceito", "Ini.Transp.", "Conclusao", "Cancelado",
        "T.Aceite(min)", "T.Desloc.(min)", "T.Transp.(min)", "T.Total(min)",
        "Motivo Cancelamento", "Obs."
    ]
    _titulo(ws, f"TODOS OS CHAMADOS — {nome_periodo} — {_data_pt(now)}", len(headers))
    for j, h in enumerate(headers, 1):
        _hdr(ws.cell(row=2, column=j, value=h))
    ws.row_dimensions[2].height = 22

    _STATUS_COR = {
        'concluido': _VERDE, 'cancelado': _VERMELHO,
        'aguardando': _HAC, 'aceito': _AZUL, 'em_transporte': _LARANJA
    }
    keys = [
        'id','tipo','nm_paciente','leito_origem','setor_origem_nome','destino_nome',
        'prioridade','status','solicitante_nome','padioleiro',
        'criado_em','dt_aceite','dt_inicio_transporte','dt_conclusao','dt_cancelamento',
        't_aceite_min','t_deslocamento_min','t_transporte_min','t_total_min',
        'motivo_cancelamento','observacao'
    ]
    for i, chamado in enumerate(todos_hoje, 3):
        zebra = i % 2 == 0
        for j, key in enumerate(keys, 1):
            v = chamado.get(key)
            cell = ws.cell(row=i, column=j, value=v)
            _borda(cell)
            cell.alignment = Alignment(vertical="center", wrap_text=(j >= 20))
            if zebra and key not in ('prioridade', 'status'):
                cell.fill = PatternFill("solid", fgColor=_ZEBRA)
            if key == 'status' and v in _STATUS_COR:
                cell.font = Font(bold=True, color=_BRANCO)
                cell.fill = PatternFill("solid", fgColor=_STATUS_COR[v])
            elif key == 'prioridade' and v == 'urgente':
                cell.font = Font(bold=True, color=_BRANCO)
                cell.fill = PatternFill("solid", fgColor=_LARANJA)

    last_row = 2 + len(todos_hoje)
    if todos_hoje:
        for col_n in range(16, 20):
            _cor_tempo(ws, get_column_letter(col_n), 3, last_row)
    _autowidth(ws)


def _aba_cancelamentos(wb, cancelados):
    ws = wb.create_sheet("Cancelamentos")
    ws.sheet_view.showGridLines = False
    headers = ["#", "Paciente", "Setor Origem", "Destino", "Padioleiro",
               "Solicitante", "H.Criacao", "H.Cancel.", "Motivo"]
    _titulo(ws, "CANCELAMENTOS DO DIA", len(headers))
    for j, h in enumerate(headers, 1):
        _hdr(ws.cell(row=2, column=j, value=h), cor=_VERMELHO)

    keys = ['id','nm_paciente','setor_origem_nome','destino_nome','padioleiro_nome',
            'solicitante_nome','hora_criacao','hora_cancelamento','motivo_cancelamento']
    for i, c in enumerate(cancelados, 3):
        zebra = i % 2 == 0
        for j, key in enumerate(keys, 1):
            cell = ws.cell(row=i, column=j, value=c.get(key))
            _borda(cell)
            if zebra: cell.fill = PatternFill("solid", fgColor="FEE2E2")
            if j == 9:
                cell.alignment = Alignment(wrap_text=True)
                cell.font = Font(color=_VERMELHO)
    _autowidth(ws)


def gerar_excel(resumo, por_padioleiro, por_setor, por_hora, tendencia_7d, todos_hoje, cancelados,
                now, horas_periodo, nome_periodo):
    wb = Workbook()
    wb.remove(wb.active)
    _aba_resumo(wb, resumo, now, nome_periodo)
    _aba_por_padioleiro(wb, por_padioleiro)
    _aba_por_setor(wb, por_setor)
    _aba_por_hora(wb, por_hora, horas_periodo, nome_periodo)
    _aba_tendencia(wb, tendencia_7d)
    _aba_todos_chamados(wb, todos_hoje, now, nome_periodo)
    if cancelados:
        _aba_cancelamentos(wb, cancelados)
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(tmp.name)
    tmp.close()
    return tmp.name

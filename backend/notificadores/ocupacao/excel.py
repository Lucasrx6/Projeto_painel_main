# -*- coding: utf-8 -*-
import tempfile
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .utils import (_CANDIDATOS_SETOR, _LABELS_NOME, _label_coluna, _nome_setor, _anonimizar, _cor_taxa)

_COR_HEADER  = "9B1C24"
_COR_TITULO  = "DC3545"
_COR_CRITICO = "C00000"
_COR_LIVRE   = "375623"
_COR_ALERTA  = "FF8C00"
_COR_BRANCO  = "FFFFFF"
_COR_ZEBRA   = "FDECEA"


def _hdr(cell, cor=_COR_HEADER):
    cell.font = Font(bold=True, color=_COR_BRANCO, size=11)
    cell.fill = PatternFill("solid", fgColor=cor)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _borda(cell)


def _borda(cell):
    s = Side(style="thin", color="AAAAAA")
    cell.border = Border(left=s, right=s, top=s, bottom=s)


def _autowidth(ws, min_w=10, max_w=50):
    for col in ws.columns:
        letra = get_column_letter(col[0].column)
        w = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[letra].width = min(max(w + 2, min_w), max_w)


def _aba_resumo(wb, dashboard, now):
    ws = wb.create_sheet("Resumo Geral")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:B1")
    t = ws["A1"]
    t.value = "OCUPACAO HOSPITALAR - HAC - {}".format(now.strftime('%d/%m/%Y %H:%M'))
    t.font = Font(bold=True, color=_COR_BRANCO, size=14)
    t.fill = PatternFill("solid", fgColor=_COR_TITULO)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.append([])

    metricas = [
        ("Indicador",                   "Valor"),
        ("Total de Leitos",             dashboard.get('total_leitos', '-')),
        ("Leitos Ocupados",             dashboard.get('leitos_ocupados', '-')),
        ("Leitos Livres",               dashboard.get('leitos_livres', '-')),
        ("Em Higienizacao",             dashboard.get('leitos_higienizacao', '-')),
        ("Interditados",                dashboard.get('leitos_interditados', '-')),
        ("Taxa de Ocupacao (%)",        dashboard.get('taxa_ocupacao_geral', '-')),
        ("Taxa de Disponibilidade (%)", dashboard.get('taxa_disponibilidade', '-') or '-'),
        ("Total de Setores",            dashboard.get('total_setores', '-')),
        ("Media de Permanencia (dias)", dashboard.get('media_permanencia_geral', '-')),
        ("Ultima Atualizacao",          str(dashboard.get('ultima_atualizacao', '-'))),
    ]

    for i, (indicador, valor) in enumerate(metricas):
        r = i + 3
        c1 = ws.cell(row=r, column=1, value=indicador)
        c2 = ws.cell(row=r, column=2, value=valor)
        if i == 0:
            _hdr(c1); _hdr(c2)
        else:
            c1.font = Font(bold=True)
            c1.fill = PatternFill("solid", fgColor="F5D0D3")
            _borda(c1); _borda(c2)
            if indicador == "Taxa de Ocupacao (%)":
                try:
                    v = float(valor)
                    cor = _COR_CRITICO if v >= 90 else (_COR_ALERTA if v >= 75 else _COR_LIVRE)
                    c2.font = Font(bold=True, color=_COR_BRANCO)
                    c2.fill = PatternFill("solid", fgColor=cor)
                except (TypeError, ValueError):
                    pass

    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 22


def _aba_por_setor(wb, cols, setores):
    ws = wb.create_sheet("Por Setor")
    ws.sheet_view.showGridLines = False
    if not cols:
        ws["A1"] = "Sem dados"; return

    for j, col in enumerate(cols, 1):
        _hdr(ws.cell(row=1, column=j, value=_label_coluna(col)))
    ws.row_dimensions[1].height = 22

    taxa_idx = next((i for i, c in enumerate(cols) if 'taxa' in c.lower() and 'ocup' in c.lower()), None)

    for i, row in enumerate(setores, 2):
        zebra = i % 2 == 0
        for j, col in enumerate(cols, 1):
            val = row.get(col)
            if col.lower() in _CANDIDATOS_SETOR:
                val = _nome_setor(row)
            cell = ws.cell(row=i, column=j, value=val)
            _borda(cell)
            cell.alignment = Alignment(vertical="center")
            if zebra:
                cell.fill = PatternFill("solid", fgColor=_COR_ZEBRA)
            if taxa_idx is not None and j == taxa_idx + 1:
                try:
                    v = float(row.get(col, 0) or 0)
                    cor = _COR_CRITICO if v >= 90 else (_COR_ALERTA if v >= 75 else None)
                    if cor:
                        cell.font = Font(bold=True, color=_COR_BRANCO)
                        cell.fill = PatternFill("solid", fgColor=cor)
                except (TypeError, ValueError):
                    pass

    _autowidth(ws)


def _aba_pacientes(wb, cols, pacientes):
    ws = wb.create_sheet("Pacientes Internados")
    ws.sheet_view.showGridLines = False
    if not cols:
        ws["A1"] = "Sem dados"; return

    for j, col in enumerate(cols, 1):
        _hdr(ws.cell(row=1, column=j, value=_label_coluna(col)))
    ws.row_dimensions[1].height = 22

    for i, row in enumerate(pacientes, 2):
        zebra = i % 2 == 0
        for j, col in enumerate(cols, 1):
            val = row.get(col)
            if isinstance(val, datetime):
                val = val.strftime('%d/%m/%Y %H:%M')
            if col.lower() in _CANDIDATOS_SETOR:
                val = _nome_setor(row)
            elif _label_coluna(col).lower() in _LABELS_NOME and val:
                val = _anonimizar(val)
            cell = ws.cell(row=i, column=j, value=val)
            _borda(cell)
            cell.alignment = Alignment(vertical="center")
            if zebra:
                cell.fill = PatternFill("solid", fgColor=_COR_ZEBRA)

    _autowidth(ws)


def gerar_excel(dashboard, cols_setor, setores, cols_pac, pacientes, now):
    wb = Workbook()
    wb.remove(wb.active)
    _aba_resumo(wb, dashboard, now)
    _aba_por_setor(wb, cols_setor, setores)
    _aba_pacientes(wb, cols_pac, pacientes)

    tmp = tempfile.NamedTemporaryFile(
        delete=False, suffix='.xlsx',
        prefix='ocupacao_hac_{}_{}_'.format(now.strftime('%Y%m%d'), now.strftime('%H%M'))
    )
    wb.save(tmp.name)
    tmp.close()
    return tmp.name

"""
Painel 36 - Gestao e Relatorios do Sistema Padioleiro
"""
from flask import Blueprint, jsonify, request, send_from_directory, send_file, session, current_app
from datetime import datetime, date
from decimal import Decimal
from backend.database import get_db_cursor
from backend.middleware.decorators import login_required, panel_permission_required
from backend.cache import cache_route
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule

painel36_bp = Blueprint('painel36', __name__)

# Whitelists para UPDATEs dinâmicos — nunca iterar sobre request diretamente
_CAMPOS_PADIOLEIRO     = ('nome', 'matricula', 'turno')
_CAMPOS_TIPO_MOVIMENTO = ('nome', 'icone', 'cor', 'ordem')
_CAMPOS_DESTINO        = ('nome', 'tipo_movimento_id', 'ordem')
_CAMPOS_ORIGEM         = ('nome', 'ordem')

# ── Constantes Excel ──────────────────────────────────────────
_X_HAC      = "9B1C24"
_X_VERDE    = "28A745"
_X_VERMELHO = "DC3545"
_X_LARANJA  = "E67E00"
_X_AZUL     = "17A2B8"
_X_BRANCO   = "FFFFFF"
_X_ZEBRA    = "FEF0F0"
_X_STATUS   = {
    'concluido': _X_VERDE, 'cancelado': _X_VERMELHO,
    'aguardando': _X_HAC,  'aceito': _X_AZUL, 'em_transporte': _X_LARANJA,
}


# ── Helpers ───────────────────────────────────────────────────

def _serial(row):
    """Converte datetime/Decimal para tipos serializáveis em JSON."""
    out = {}
    for k, v in row.items():
        if isinstance(v, datetime):
            out[k] = v.isoformat()
        elif isinstance(v, Decimal):
            out[k] = float(v)
        else:
            out[k] = v
    return out


def _periodo_where(req):
    """Retorna (where_list, params) para filtro de período."""
    data_inicio = req.args.get('data_inicio', '').strip()
    data_fim    = req.args.get('data_fim', '').strip()
    if data_inicio and data_fim:
        return (
            ["criado_em >= %s::date", "criado_em < (%s::date + INTERVAL '1 day')"],
            [data_inicio, data_fim],
        )
    dias = min(int(req.args.get('dias', 30)), 365)
    return (["criado_em >= NOW() - (%s || ' days')::INTERVAL"], [str(dias)])


def _cfg_atualizar(tabela, campos_wl, dados, rec_id, mensagem, log_tag, strip_str=False):
    """UPDATE genérico para tabelas de configuração com whitelist de colunas."""
    try:
        with get_db_cursor(use_dict_cursor=False) as cursor:
            fields, params = [], []
            for campo in campos_wl:
                if campo not in dados:
                    continue
                val = dados[campo]
                if strip_str and isinstance(val, str):
                    val = val.strip() or None
                if strip_str and campo == 'nome' and not val:
                    return jsonify({'success': False, 'error': 'Nome nao pode ser vazio'}), 400
                fields.append(f'{campo} = %s')
                params.append(val)
            if 'ativo' in dados:
                fields.append('ativo = %s')
                params.append(bool(dados['ativo']))
            if not fields:
                return jsonify({'success': False, 'error': 'Nada para atualizar'}), 400
            fields.append('atualizado_em = NOW()')
            params.append(rec_id)
            cursor.execute(f"UPDATE {tabela} SET {', '.join(fields)} WHERE id = %s", params)
        return jsonify({'success': True, 'message': mensagem})
    except Exception as e:
        current_app.logger.error(f'Erro {log_tag}: {e}', exc_info=True)
        return jsonify({'success': False, 'error': 'Erro ao atualizar'}), 500


# ── Helpers Excel ─────────────────────────────────────────────

def _x_hdr(cell, cor=_X_HAC):
    cell.font      = Font(bold=True, color=_X_BRANCO, size=11)
    cell.fill      = PatternFill("solid", fgColor=cor)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _x_borda(cell)

def _x_borda(cell):
    s = Side(style="thin", color="CCCCCC")
    cell.border = Border(left=s, right=s, top=s, bottom=s)

def _x_autowidth(ws, min_w=10, max_w=45):
    for col in ws.columns:
        letra = get_column_letter(col[0].column)
        w = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[letra].width = min(max(w + 3, min_w), max_w)

def _x_titulo(ws, texto, ncols, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=texto)
    c.font      = Font(bold=True, color=_X_BRANCO, size=13)
    c.fill      = PatternFill("solid", fgColor=_X_HAC)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28

def _x_cor_tempo(ws, col_letra, row_ini, row_fim):
    ws.conditional_formatting.add(
        f"{col_letra}{row_ini}:{col_letra}{row_fim}",
        ColorScaleRule(
            start_type='num', start_value=0,  start_color="63BE7B",
            mid_type='num',   mid_value=20,   mid_color="FFEB84",
            end_type='num',   end_value=60,   end_color="F8696B",
        ),
    )


# =========================================================
# PÁGINA
# =========================================================

@painel36_bp.route('/painel/painel36')
@login_required
@panel_permission_required('painel36')
def painel36():
    return send_from_directory('paineis/painel36', 'index.html')


# =========================================================
# DASHBOARD
# =========================================================

@painel36_bp.route('/api/paineis/painel36/dashboard')
@login_required
@panel_permission_required('painel36')
@cache_route(ttl=30, key_prefix='painel36:dashboard', vary_by_user=False)
def api_painel36_dashboard():
    try:
        with get_db_cursor() as cursor:
            cursor.execute("""
                SELECT
                    COUNT(*) FILTER (WHERE status = 'aguardando')                              AS aguardando,
                    COUNT(*) FILTER (WHERE status = 'aceito')                                  AS aceito,
                    COUNT(*) FILTER (WHERE status = 'em_transporte')                           AS em_transporte,
                    COUNT(*) FILTER (WHERE status = 'concluido'  AND criado_em >= CURRENT_DATE) AS concluidos_hoje,
                    COUNT(*) FILTER (WHERE status = 'cancelado'  AND criado_em >= CURRENT_DATE) AS cancelados_hoje,
                    COUNT(*) FILTER (WHERE criado_em >= CURRENT_DATE)                          AS total_hoje,
                    COUNT(*) FILTER (WHERE prioridade = 'urgente' AND status = 'aguardando')   AS urgentes_aguardando,
                    ROUND(AVG(EXTRACT(EPOCH FROM (dt_conclusao - criado_em)) / 60)
                        FILTER (WHERE status = 'concluido' AND criado_em >= CURRENT_DATE
                                AND dt_conclusao IS NOT NULL
                                AND EXTRACT(EPOCH FROM (dt_conclusao - criado_em)) / 60 <= 300), 1) AS tempo_medio_total_hoje,
                    ROUND(AVG(EXTRACT(EPOCH FROM (dt_aceite - criado_em)) / 60)
                        FILTER (WHERE dt_aceite IS NOT NULL AND criado_em >= CURRENT_DATE
                                AND EXTRACT(EPOCH FROM (dt_aceite - criado_em)) / 60 <= 300), 1)    AS tempo_medio_aceite_hoje,
                    ROUND(AVG(EXTRACT(EPOCH FROM (dt_inicio_transporte - dt_aceite)) / 60)
                        FILTER (WHERE dt_inicio_transporte IS NOT NULL AND dt_aceite IS NOT NULL
                                AND criado_em >= CURRENT_DATE
                                AND EXTRACT(EPOCH FROM (dt_inicio_transporte - dt_aceite)) / 60 <= 300), 1) AS tempo_medio_deslocamento_hoje,
                    ROUND(AVG(EXTRACT(EPOCH FROM (dt_conclusao - dt_inicio_transporte)) / 60)
                        FILTER (WHERE status = 'concluido' AND dt_inicio_transporte IS NOT NULL
                                AND dt_conclusao IS NOT NULL AND criado_em >= CURRENT_DATE
                                AND EXTRACT(EPOCH FROM (dt_conclusao - criado_em)) / 60 <= 300), 1) AS tempo_medio_transporte_hoje
                FROM padioleiro_chamados
                WHERE status IN ('aguardando', 'aceito', 'em_transporte')
                   OR criado_em >= CURRENT_DATE
            """)
            stats = dict(cursor.fetchone() or {})
            for k, v in stats.items():
                if v is not None and hasattr(v, '__float__'):
                    stats[k] = float(v)

            cursor.execute("""
                SELECT
                    id, tipo_movimento_nome, nm_paciente, nr_atendimento,
                    leito_origem, setor_origem_nome, destino_nome,
                    prioridade, status, solicitante_nome, padioleiro_nome,
                    criado_em, dt_aceite, dt_inicio_transporte,
                    ROUND(EXTRACT(EPOCH FROM (NOW() - criado_em)) / 60, 1) AS minutos_espera
                FROM padioleiro_chamados
                WHERE status IN ('aguardando', 'aceito', 'em_transporte')
                ORDER BY
                    CASE prioridade WHEN 'urgente' THEN 0 ELSE 1 END,
                    criado_em ASC
            """)
            ativos = []
            for row in cursor.fetchall():
                c = _serial(dict(row))
                if c.get('minutos_espera') is not None:
                    c['minutos_espera'] = float(c['minutos_espera'])
                ativos.append(c)

        return jsonify({'success': True, 'stats': stats, 'ativos': ativos,
                        'timestamp': datetime.now().isoformat()})
    except Exception as e:
        current_app.logger.error(f'Erro dashboard painel36: {e}', exc_info=True)
        return jsonify({'success': False, 'error': 'Erro ao buscar dados'}), 500


# =========================================================
# LISTAGEM COM FILTROS
# =========================================================

@painel36_bp.route('/api/paineis/painel36/chamados')
@login_required
@panel_permission_required('painel36')
@cache_route(ttl=30, key_prefix='painel36:chamados', vary_by_user=False, vary_by_query=True)
def api_painel36_chamados():
    setor         = request.args.get('setor', '').strip()
    padioleiro_id = request.args.get('padioleiro_id', '').strip()
    tipo_id       = request.args.get('tipo_id', '').strip()
    status        = request.args.get('status', '').strip()
    prioridade    = request.args.get('prioridade', '').strip()

    try:
        with get_db_cursor() as cursor:
            where, params = _periodo_where(request)
            if setor:         where.append("setor_origem_nome ILIKE %s"); params.append(f'%{setor}%')
            if padioleiro_id: where.append("padioleiro_id = %s");         params.append(padioleiro_id)
            if tipo_id:       where.append("tipo_movimento_id = %s");     params.append(tipo_id)
            if status:        where.append("status = %s");                params.append(status)
            if prioridade:    where.append("prioridade = %s");            params.append(prioridade)

            cursor.execute("""
                SELECT
                    id, tipo_movimento_nome, nm_paciente, nr_atendimento,
                    leito_origem, setor_origem_nome, destino_nome, destino_complemento,
                    prioridade, status, solicitante_nome, padioleiro_nome, observacao,
                    criado_em, dt_aceite, dt_inicio_transporte, dt_conclusao, dt_cancelamento,
                    motivo_cancelamento,
                    CASE WHEN dt_aceite IS NOT NULL
                         THEN ROUND(EXTRACT(EPOCH FROM (dt_aceite - criado_em)) / 60, 1)
                    END AS tempo_aceite_min,
                    CASE WHEN dt_inicio_transporte IS NOT NULL AND dt_aceite IS NOT NULL
                         THEN ROUND(EXTRACT(EPOCH FROM (dt_inicio_transporte - dt_aceite)) / 60, 1)
                    END AS tempo_deslocamento_min,
                    CASE WHEN dt_conclusao IS NOT NULL AND dt_inicio_transporte IS NOT NULL
                         THEN ROUND(EXTRACT(EPOCH FROM (dt_conclusao - dt_inicio_transporte)) / 60, 1)
                    END AS tempo_transporte_min,
                    CASE WHEN dt_conclusao IS NOT NULL
                         THEN ROUND(EXTRACT(EPOCH FROM (dt_conclusao - criado_em)) / 60, 1)
                    END AS tempo_total_min
                FROM padioleiro_chamados
                WHERE """ + ' AND '.join(where) + """
                ORDER BY criado_em DESC
                LIMIT 500
            """, params)

            chamados = [_serial(dict(row)) for row in cursor.fetchall()]
            return jsonify({'success': True, 'chamados': chamados, 'total': len(chamados)})

    except Exception as e:
        current_app.logger.error(f'Erro chamados painel36: {e}', exc_info=True)
        return jsonify({'success': False, 'error': 'Erro ao buscar chamados'}), 500


# =========================================================
# CANCELAR CHAMADO (Gestão)
# =========================================================

@painel36_bp.route('/api/paineis/painel36/chamados/<int:chamado_id>/cancelar', methods=['PUT'])
@login_required
@panel_permission_required('painel36')
def api_painel36_cancelar(chamado_id):
    motivo = ((request.get_json() or {}).get('motivo') or '').strip()
    if len(motivo) < 10:
        return jsonify({'success': False,
                        'error': 'O motivo do cancelamento deve ter pelo menos 10 caracteres'}), 400
    try:
        with get_db_cursor() as cursor:
            cursor.execute("SELECT status FROM padioleiro_chamados WHERE id = %s", (chamado_id,))
            chamado = cursor.fetchone()
            if not chamado:
                return jsonify({'success': False, 'error': 'Chamado nao encontrado'}), 404
            if chamado['status'] in ('concluido', 'cancelado'):
                return jsonify({'success': False,
                                'error': f'Chamado nao pode ser cancelado no status atual: {chamado["status"]}'}), 400
            cursor.execute("""
                UPDATE padioleiro_chamados
                SET status = 'cancelado', dt_cancelamento = NOW(),
                    motivo_cancelamento = %s, atualizado_em = NOW()
                WHERE id = %s
            """, (f"[Cancelado pela Gestão] {motivo}", chamado_id))
        return jsonify({'success': True, 'message': 'Chamado cancelado administrativamente com sucesso'})
    except Exception as e:
        current_app.logger.error(f'Erro cancelar painel36: {e}', exc_info=True)
        return jsonify({'success': False, 'error': 'Erro ao cancelar chamado'}), 500


# =========================================================
# ANALYTICS POR SETOR
# =========================================================

@painel36_bp.route('/api/paineis/painel36/por-setor')
@login_required
@panel_permission_required('painel36')
@cache_route(ttl=120, key_prefix='painel36:por-setor', vary_by_user=False, vary_by_query=True)
def api_painel36_por_setor():
    try:
        with get_db_cursor() as cursor:
            where, params = _periodo_where(request)
            cursor.execute("""
                SELECT
                    setor_origem_nome                                                           AS setor,
                    COUNT(*)                                                                    AS total,
                    COUNT(*) FILTER (WHERE status = 'concluido')                               AS concluidos,
                    COUNT(*) FILTER (WHERE status = 'cancelado')                               AS cancelados,
                    COUNT(*) FILTER (WHERE prioridade = 'urgente')                             AS urgentes,
                    ROUND(AVG(EXTRACT(EPOCH FROM (dt_aceite - criado_em)) / 60)
                        FILTER (WHERE dt_aceite IS NOT NULL), 1)                               AS tempo_medio_aceite_min,
                    ROUND(AVG(EXTRACT(EPOCH FROM (dt_inicio_transporte - dt_aceite)) / 60)
                        FILTER (WHERE dt_inicio_transporte IS NOT NULL AND dt_aceite IS NOT NULL), 1) AS tempo_medio_deslocamento_min,
                    ROUND(AVG(EXTRACT(EPOCH FROM (dt_conclusao - criado_em)) / 60)
                        FILTER (WHERE status = 'concluido' AND dt_conclusao IS NOT NULL), 1)   AS tempo_medio_total_min
                FROM padioleiro_chamados
                WHERE """ + ' AND '.join(where) + """
                GROUP BY setor_origem_nome
                ORDER BY total DESC
            """, params)
            setores = [_serial(dict(row)) for row in cursor.fetchall()]
        return jsonify({'success': True, 'setores': setores})
    except Exception as e:
        current_app.logger.error(f'Erro por-setor painel36: {e}', exc_info=True)
        return jsonify({'success': False, 'error': 'Erro ao buscar dados'}), 500


# =========================================================
# ANALYTICS POR PADIOLEIRO
# =========================================================

@painel36_bp.route('/api/paineis/painel36/por-padioleiro')
@login_required
@panel_permission_required('painel36')
@cache_route(ttl=120, key_prefix='painel36:por-padioleiro', vary_by_user=False, vary_by_query=True)
def api_painel36_por_padioleiro():
    try:
        with get_db_cursor() as cursor:
            where, params = _periodo_where(request)
            where.append("padioleiro_nome IS NOT NULL")
            cursor.execute("""
                SELECT
                    padioleiro_nome                                                             AS padioleiro,
                    COUNT(*)                                                                    AS total,
                    COUNT(*) FILTER (WHERE status = 'concluido')                               AS concluidos,
                    COUNT(*) FILTER (WHERE status = 'cancelado')                               AS cancelados,
                    COUNT(*) FILTER (WHERE prioridade = 'urgente')                             AS urgentes,
                    ROUND(AVG(EXTRACT(EPOCH FROM (dt_aceite - criado_em)) / 60)
                        FILTER (WHERE dt_aceite IS NOT NULL), 1)                               AS tempo_medio_aceite_min,
                    ROUND(AVG(EXTRACT(EPOCH FROM (dt_inicio_transporte - dt_aceite)) / 60)
                        FILTER (WHERE dt_inicio_transporte IS NOT NULL AND dt_aceite IS NOT NULL), 1) AS tempo_medio_deslocamento_min,
                    ROUND(AVG(EXTRACT(EPOCH FROM (dt_conclusao - dt_inicio_transporte)) / 60)
                        FILTER (WHERE status = 'concluido' AND dt_conclusao IS NOT NULL
                                AND dt_inicio_transporte IS NOT NULL), 1)                      AS tempo_medio_transporte_min,
                    ROUND(AVG(EXTRACT(EPOCH FROM (dt_conclusao - criado_em)) / 60)
                        FILTER (WHERE status = 'concluido' AND dt_conclusao IS NOT NULL), 1)   AS tempo_medio_total_min
                FROM padioleiro_chamados
                WHERE """ + ' AND '.join(where) + """
                GROUP BY padioleiro_nome
                ORDER BY concluidos DESC
            """, params)
            padioleiros = [_serial(dict(row)) for row in cursor.fetchall()]
        return jsonify({'success': True, 'padioleiros': padioleiros})
    except Exception as e:
        current_app.logger.error(f'Erro por-padioleiro painel36: {e}', exc_info=True)
        return jsonify({'success': False, 'error': 'Erro ao buscar dados'}), 500


# =========================================================
# EXPORTAR EXCEL
# =========================================================

@painel36_bp.route('/api/paineis/painel36/exportar')
@login_required
@panel_permission_required('painel36')
def api_painel36_exportar():
    status      = request.args.get('status', '').strip()
    prioridade  = request.args.get('prioridade', '').strip()
    setor       = request.args.get('setor', '').strip()
    data_inicio = request.args.get('data_inicio', '').strip()
    data_fim    = request.args.get('data_fim', '').strip()

    try:
        with get_db_cursor() as cursor:
            where, params = _periodo_where(request)
            if status:    where.append("status = %s");                params.append(status)
            if prioridade: where.append("prioridade = %s");           params.append(prioridade)
            if setor:     where.append("setor_origem_nome ILIKE %s"); params.append(f'%{setor}%')
            where_sql = ' AND '.join(where)

            cursor.execute(f"""
                SELECT
                    id, tipo_movimento_nome, nm_paciente, nr_atendimento,
                    leito_origem, setor_origem_nome, destino_nome, destino_complemento,
                    prioridade, status, solicitante_nome, padioleiro_nome, observacao,
                    TO_CHAR(criado_em,            'DD/MM/YYYY HH24:MI') AS criado_em,
                    TO_CHAR(dt_aceite,            'DD/MM/YYYY HH24:MI') AS dt_aceite,
                    TO_CHAR(dt_inicio_transporte, 'DD/MM/YYYY HH24:MI') AS dt_inicio_transporte,
                    TO_CHAR(dt_conclusao,         'DD/MM/YYYY HH24:MI') AS dt_conclusao,
                    TO_CHAR(dt_cancelamento,      'DD/MM/YYYY HH24:MI') AS dt_cancelamento,
                    motivo_cancelamento,
                    CASE WHEN dt_aceite IS NOT NULL
                         THEN ROUND(EXTRACT(EPOCH FROM (dt_aceite - criado_em)) / 60, 1) END AS t_aceite_min,
                    CASE WHEN dt_inicio_transporte IS NOT NULL AND dt_aceite IS NOT NULL
                         THEN ROUND(EXTRACT(EPOCH FROM (dt_inicio_transporte - dt_aceite)) / 60, 1) END AS t_deslocamento_min,
                    CASE WHEN dt_conclusao IS NOT NULL AND dt_inicio_transporte IS NOT NULL
                         THEN ROUND(EXTRACT(EPOCH FROM (dt_conclusao - dt_inicio_transporte)) / 60, 1) END AS t_transporte_min,
                    CASE WHEN dt_conclusao IS NOT NULL
                         THEN ROUND(EXTRACT(EPOCH FROM (dt_conclusao - criado_em)) / 60, 1)
                         WHEN status = 'cancelado' AND dt_cancelamento IS NOT NULL
                         THEN ROUND(EXTRACT(EPOCH FROM (dt_cancelamento - criado_em)) / 60, 1)
                    END AS t_total_min
                FROM padioleiro_chamados WHERE {where_sql} ORDER BY criado_em DESC
            """, params)
            chamados = [dict(r) for r in cursor.fetchall()]

            cursor.execute(f"""
                SELECT
                    COALESCE(padioleiro_nome,'(não atribuído)') AS padioleiro,
                    COUNT(*) AS total,
                    COUNT(*) FILTER (WHERE status='concluido')   AS concluidos,
                    COUNT(*) FILTER (WHERE status='cancelado')   AS cancelados,
                    COUNT(*) FILTER (WHERE prioridade='urgente') AS urgentes,
                    ROUND(AVG(CASE WHEN dt_aceite IS NOT NULL
                        THEN EXTRACT(EPOCH FROM (dt_aceite - criado_em))/60 END)::numeric,1) AS media_aceite_min,
                    ROUND(AVG(CASE WHEN dt_inicio_transporte IS NOT NULL AND dt_aceite IS NOT NULL
                        THEN EXTRACT(EPOCH FROM (dt_inicio_transporte - dt_aceite))/60 END)::numeric,1) AS media_deslocamento_min,
                    ROUND(AVG(CASE WHEN dt_conclusao IS NOT NULL AND dt_inicio_transporte IS NOT NULL
                        THEN EXTRACT(EPOCH FROM (dt_conclusao - dt_inicio_transporte))/60 END)::numeric,1) AS media_transporte_min,
                    ROUND(AVG(CASE WHEN dt_conclusao IS NOT NULL
                        THEN EXTRACT(EPOCH FROM (dt_conclusao - criado_em))/60 END)::numeric,1) AS media_total_min
                FROM padioleiro_chamados WHERE {where_sql}
                GROUP BY padioleiro_nome ORDER BY total DESC
            """, params)
            por_padioleiro = [dict(r) for r in cursor.fetchall()]

            cursor.execute(f"""
                SELECT
                    COALESCE(setor_origem_nome,'(sem setor)') AS setor,
                    COUNT(*) AS total,
                    COUNT(*) FILTER (WHERE status='concluido')   AS concluidos,
                    COUNT(*) FILTER (WHERE status='cancelado')   AS cancelados,
                    COUNT(*) FILTER (WHERE prioridade='urgente') AS urgentes
                FROM padioleiro_chamados WHERE {where_sql}
                GROUP BY setor_origem_nome ORDER BY total DESC LIMIT 30
            """, params)
            por_setor = [dict(r) for r in cursor.fetchall()]

        # ── Montar filtros_txt ────────────────────────────────
        now = datetime.now()
        if data_inicio and data_fim:
            di = datetime.strptime(data_inicio, '%Y-%m-%d').strftime('%d/%m/%Y')
            df = datetime.strptime(data_fim,    '%Y-%m-%d').strftime('%d/%m/%Y')
            filtros_txt = f"Período: {di} a {df}"
        else:
            filtros_txt = f"Últimos {int(request.args.get('dias', 30))} dia(s)"
        if status:    filtros_txt += f"  |  Status: {status}"
        if prioridade: filtros_txt += f"  |  Prioridade: {prioridade}"
        if setor:     filtros_txt += f"  |  Setor: {setor}"

        # ── Workbook ──────────────────────────────────────────
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Aba 1: Chamados
        ws1 = wb.create_sheet("Chamados")
        ws1.sheet_view.showGridLines = False
        ws1.freeze_panes = "A3"
        _x_titulo(ws1, f"CHAMADOS PADIOLEIRO — HAC — {now.strftime('%d/%m/%Y %H:%M')}  |  {filtros_txt}", 23)

        hdrs1 = [
            "#", "Tipo Movimento", "Paciente", "Atendimento", "Leito Origem",
            "Setor Origem", "Destino", "Compl. Destino", "Prioridade", "Status",
            "Solicitante", "Padioleiro", "Observação",
            "Criado Em", "Aceito Em", "Ini. Transporte", "Conclusão", "Cancelado Em",
            "Motivo Cancelamento", "T.Aceite(min)", "T.Desloc.(min)", "T.Transp.(min)", "T.Total(min)",
        ]
        for j, h in enumerate(hdrs1, 1):
            _x_hdr(ws1.cell(row=2, column=j, value=h))
        ws1.row_dimensions[2].height = 22

        keys1 = [
            'id','tipo_movimento_nome','nm_paciente','nr_atendimento','leito_origem',
            'setor_origem_nome','destino_nome','destino_complemento',
            'prioridade','status','solicitante_nome','padioleiro_nome','observacao',
            'criado_em','dt_aceite','dt_inicio_transporte','dt_conclusao','dt_cancelamento',
            'motivo_cancelamento','t_aceite_min','t_deslocamento_min','t_transporte_min','t_total_min',
        ]
        for i, row in enumerate(chamados, 3):
            zebra = i % 2 == 0
            for j, key in enumerate(keys1, 1):
                v    = row.get(key)
                cell = ws1.cell(row=i, column=j, value=v)
                _x_borda(cell)
                cell.alignment = Alignment(vertical="center", wrap_text=(j in (13, 19)))
                if zebra and key not in ('prioridade', 'status'):
                    cell.fill = PatternFill("solid", fgColor=_X_ZEBRA)
                if key == 'status' and v in _X_STATUS:
                    cell.font = Font(bold=True, color=_X_BRANCO)
                    cell.fill = PatternFill("solid", fgColor=_X_STATUS[v])
                elif key == 'prioridade' and v == 'urgente':
                    cell.font = Font(bold=True, color=_X_BRANCO)
                    cell.fill = PatternFill("solid", fgColor=_X_LARANJA)

        last1 = 2 + len(chamados)
        if chamados:
            for col_n in range(20, 24):
                _x_cor_tempo(ws1, get_column_letter(col_n), 3, last1)
        _x_autowidth(ws1)

        # Aba 2: Por Padioleiro
        ws2 = wb.create_sheet("Por Padioleiro")
        ws2.sheet_view.showGridLines = False
        hdrs2 = ["Padioleiro","Total","Concluídos","Cancelados","Urgentes",
                 "T.Aceite(min)","T.Desloc.(min)","T.Transp.(min)","T.Total(min)"]
        _x_titulo(ws2, f"POR PADIOLEIRO — {filtros_txt}", len(hdrs2))
        for j, h in enumerate(hdrs2, 1):
            _x_hdr(ws2.cell(row=2, column=j, value=h))

        pad_cols = [
            ('padioleiro',None),('total',None),('concluidos',_X_VERDE),
            ('cancelados',_X_VERMELHO),('urgentes',_X_LARANJA),
            ('media_aceite_min',None),('media_deslocamento_min',None),
            ('media_transporte_min',None),('media_total_min',None),
        ]
        for i, p in enumerate(por_padioleiro, 3):
            zebra = i % 2 == 0
            for j, (key, cor_txt) in enumerate(pad_cols, 1):
                v    = p.get(key)
                cell = ws2.cell(row=i, column=j, value=float(v) if v is not None and j > 5 else v)
                _x_borda(cell)
                if zebra:    cell.fill = PatternFill("solid", fgColor=_X_ZEBRA)
                if j == 1:   cell.font = Font(bold=True)
                elif cor_txt: cell.font = Font(bold=True, color=cor_txt)

        last2 = 2 + len(por_padioleiro)
        if por_padioleiro:
            for col_l in ['F', 'G', 'H', 'I']:
                _x_cor_tempo(ws2, col_l, 3, last2)
            c = BarChart()
            c.type = "bar"; c.grouping = "clustered"
            c.title = "Movimentos por Padioleiro"
            c.style = 10; c.height = 12; c.width = 22
            c.x_axis.title = "Quantidade"
            c.add_data(Reference(ws2, min_col=2, max_col=5, min_row=2, max_row=last2), titles_from_data=True)
            c.set_categories(Reference(ws2, min_col=1, min_row=3, max_row=last2))
            for idx, cor in enumerate([_X_HAC, _X_VERDE, _X_VERMELHO, _X_LARANJA]):
                if idx < len(c.series):
                    c.series[idx].graphicalProperties.solidFill = cor
            ws2.add_chart(c, f"A{last2 + 3}")
        _x_autowidth(ws2)

        # Aba 3: Por Setor
        ws3 = wb.create_sheet("Por Setor")
        ws3.sheet_view.showGridLines = False
        hdrs3 = ["Setor","Total","Concluídos","Cancelados","Urgentes"]
        _x_titulo(ws3, f"POR SETOR — {filtros_txt}", len(hdrs3))
        for j, h in enumerate(hdrs3, 1):
            _x_hdr(ws3.cell(row=2, column=j, value=h))

        set_cols = [
            ('setor',None),('total',None),('concluidos',_X_VERDE),
            ('cancelados',_X_VERMELHO),('urgentes',_X_LARANJA),
        ]
        for i, s in enumerate(por_setor, 3):
            zebra = i % 2 == 0
            for j, (key, cor_txt) in enumerate(set_cols, 1):
                cell = ws3.cell(row=i, column=j, value=s.get(key))
                _x_borda(cell)
                if zebra:    cell.fill = PatternFill("solid", fgColor=_X_ZEBRA)
                if j == 1:   cell.font = Font(bold=True)
                elif cor_txt: cell.font = Font(bold=True, color=cor_txt)

        last3 = 2 + len(por_setor)
        if por_setor:
            c2 = BarChart()
            c2.type = "bar"; c2.grouping = "clustered"
            c2.title = "Chamados por Setor"
            c2.style = 10; c2.height = max(12, len(por_setor) * 0.9); c2.width = 22
            c2.x_axis.title = "Quantidade"
            c2.add_data(Reference(ws3, min_col=2, max_col=5, min_row=2, max_row=last3), titles_from_data=True)
            c2.set_categories(Reference(ws3, min_col=1, min_row=3, max_row=last3))
            for idx, cor in enumerate([_X_HAC, _X_VERDE, _X_VERMELHO, _X_LARANJA]):
                if idx < len(c2.series):
                    c2.series[idx].graphicalProperties.solidFill = cor
            ws3.add_chart(c2, f"A{last3 + 3}")
        _x_autowidth(ws3)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'chamados_padioleiro_{date.today().strftime("%Y%m%d")}.xlsx',
        )

    except Exception as e:
        current_app.logger.error(f'Erro exportar painel36: {e}', exc_info=True)
        return jsonify({'success': False, 'error': 'Erro ao exportar'}), 500


# =========================================================
# CONFIG: PADIOLEIROS
# =========================================================

@painel36_bp.route('/api/paineis/painel36/config/padioleiros')
@login_required
@panel_permission_required('painel36')
def api_painel36_cfg_pad_listar():
    try:
        with get_db_cursor() as cursor:
            cursor.execute("""
                SELECT id, nome, matricula, turno, ativo,
                       TO_CHAR(criado_em, 'DD/MM/YYYY') AS criado_em
                FROM padioleiro_cadastros ORDER BY nome
            """)
            return jsonify({'success': True, 'padioleiros': [dict(r) for r in cursor.fetchall()]})
    except Exception as e:
        current_app.logger.error(f'Erro listar padioleiros painel36: {e}', exc_info=True)
        return jsonify({'success': False, 'error': 'Erro ao buscar padioleiros'}), 500


@painel36_bp.route('/api/paineis/painel36/config/padioleiros', methods=['POST'])
@login_required
@panel_permission_required('painel36')
def api_painel36_cfg_pad_criar():
    dados = request.get_json() or {}
    nome  = (dados.get('nome') or '').strip()
    if not nome:
        return jsonify({'success': False, 'error': 'Nome e obrigatorio'}), 400
    try:
        with get_db_cursor() as cursor:
            cursor.execute("""
                INSERT INTO padioleiro_cadastros (nome, matricula, turno, ativo, criado_em)
                VALUES (%s, %s, %s, TRUE, NOW()) RETURNING id
            """, (nome, (dados.get('matricula') or '').strip() or None, dados.get('turno', 'todos')))
            return jsonify({'success': True, 'id': cursor.fetchone()['id'],
                            'message': 'Padioleiro cadastrado'}), 201
    except Exception as e:
        current_app.logger.error(f'Erro criar padioleiro painel36: {e}', exc_info=True)
        return jsonify({'success': False, 'error': 'Erro ao cadastrar padioleiro'}), 500


@painel36_bp.route('/api/paineis/painel36/config/padioleiros/<int:padioleiro_id>', methods=['PUT'])
@login_required
@panel_permission_required('painel36')
def api_painel36_cfg_pad_atualizar(padioleiro_id):
    return _cfg_atualizar(
        'padioleiro_cadastros', _CAMPOS_PADIOLEIRO,
        request.get_json() or {}, padioleiro_id,
        'Padioleiro atualizado', 'atualizar padioleiro painel36', strip_str=True,
    )


# =========================================================
# CONFIG: TIPOS DE MOVIMENTO
# =========================================================

@painel36_bp.route('/api/paineis/painel36/config/tipos-movimento')
@login_required
@panel_permission_required('painel36')
def api_painel36_cfg_tipos_listar():
    try:
        with get_db_cursor() as cursor:
            cursor.execute("SELECT id, nome, icone, cor, ativo, ordem FROM padioleiro_tipos_movimento ORDER BY ordem, nome")
            return jsonify({'success': True, 'tipos': [dict(r) for r in cursor.fetchall()]})
    except Exception as e:
        current_app.logger.error(f'Erro listar tipos painel36: {e}', exc_info=True)
        return jsonify({'success': False, 'error': 'Erro ao buscar tipos'}), 500


@painel36_bp.route('/api/paineis/painel36/config/tipos-movimento', methods=['POST'])
@login_required
@panel_permission_required('painel36')
def api_painel36_cfg_tipos_criar():
    dados = request.get_json() or {}
    nome  = (dados.get('nome') or '').strip()
    if not nome:
        return jsonify({'success': False, 'error': 'Nome e obrigatorio'}), 400
    try:
        with get_db_cursor() as cursor:
            cursor.execute("""
                INSERT INTO padioleiro_tipos_movimento (nome, icone, cor, ativo, ordem)
                VALUES (%s, %s, %s, TRUE, %s) RETURNING id
            """, (nome, dados.get('icone', 'fa-ambulance'), dados.get('cor', '#dc3545'), dados.get('ordem', 0)))
            return jsonify({'success': True, 'id': cursor.fetchone()['id']}), 201
    except Exception as e:
        current_app.logger.error(f'Erro criar tipo painel36: {e}', exc_info=True)
        return jsonify({'success': False, 'error': 'Erro ao criar tipo'}), 500


@painel36_bp.route('/api/paineis/painel36/config/tipos-movimento/<int:tipo_id>', methods=['PUT'])
@login_required
@panel_permission_required('painel36')
def api_painel36_cfg_tipos_atualizar(tipo_id):
    return _cfg_atualizar(
        'padioleiro_tipos_movimento', _CAMPOS_TIPO_MOVIMENTO,
        request.get_json() or {}, tipo_id,
        'Tipo atualizado', 'atualizar tipo painel36',
    )


# =========================================================
# CONFIG: DESTINOS
# =========================================================

@painel36_bp.route('/api/paineis/painel36/config/destinos')
@login_required
@panel_permission_required('painel36')
def api_painel36_cfg_dest_listar():
    tipo_id = request.args.get('tipo_id', '').strip()
    try:
        with get_db_cursor() as cursor:
            if tipo_id:
                cursor.execute("""
                    SELECT d.id, d.nome, d.tipo_movimento_id, t.nome AS tipo_nome, d.ativo, d.ordem
                    FROM padioleiro_destinos d
                    JOIN padioleiro_tipos_movimento t ON t.id = d.tipo_movimento_id
                    WHERE d.tipo_movimento_id = %s ORDER BY d.ordem, d.nome
                """, (tipo_id,))
            else:
                cursor.execute("""
                    SELECT d.id, d.nome, d.tipo_movimento_id, t.nome AS tipo_nome, d.ativo, d.ordem
                    FROM padioleiro_destinos d
                    JOIN padioleiro_tipos_movimento t ON t.id = d.tipo_movimento_id
                    ORDER BY t.ordem, d.ordem, d.nome
                """)
            return jsonify({'success': True, 'destinos': [dict(r) for r in cursor.fetchall()]})
    except Exception as e:
        current_app.logger.error(f'Erro listar destinos painel36: {e}', exc_info=True)
        return jsonify({'success': False, 'error': 'Erro ao buscar destinos'}), 500


@painel36_bp.route('/api/paineis/painel36/config/destinos', methods=['POST'])
@login_required
@panel_permission_required('painel36')
def api_painel36_cfg_dest_criar():
    dados             = request.get_json() or {}
    nome              = (dados.get('nome') or '').strip()
    tipo_movimento_id = dados.get('tipo_movimento_id')
    if not nome or not tipo_movimento_id:
        return jsonify({'success': False, 'error': 'Nome e tipo de movimento sao obrigatorios'}), 400
    try:
        with get_db_cursor() as cursor:
            cursor.execute("""
                INSERT INTO padioleiro_destinos (nome, tipo_movimento_id, ativo, ordem)
                VALUES (%s, %s, TRUE, %s) RETURNING id
            """, (nome, tipo_movimento_id, dados.get('ordem', 0)))
            return jsonify({'success': True, 'id': cursor.fetchone()['id']}), 201
    except Exception as e:
        current_app.logger.error(f'Erro criar destino painel36: {e}', exc_info=True)
        return jsonify({'success': False, 'error': 'Erro ao criar destino'}), 500


@painel36_bp.route('/api/paineis/painel36/config/destinos/<int:destino_id>', methods=['PUT'])
@login_required
@panel_permission_required('painel36')
def api_painel36_cfg_dest_atualizar(destino_id):
    return _cfg_atualizar(
        'padioleiro_destinos', _CAMPOS_DESTINO,
        request.get_json() or {}, destino_id,
        'Destino atualizado', 'atualizar destino painel36',
    )


# =========================================================
# CONFIG: ORIGENS
# =========================================================

@painel36_bp.route('/api/paineis/painel36/config/origens')
@login_required
@panel_permission_required('painel36')
def api_painel36_cfg_orig_listar():
    try:
        with get_db_cursor() as cursor:
            cursor.execute("SELECT id, nome, ativo, ordem FROM padioleiro_origens ORDER BY ordem, nome")
            return jsonify({'success': True, 'origens': [dict(r) for r in cursor.fetchall()]})
    except Exception as e:
        current_app.logger.error(f'Erro listar origens painel36: {e}', exc_info=True)
        return jsonify({'success': False, 'error': 'Erro ao buscar origens'}), 500


@painel36_bp.route('/api/paineis/painel36/config/origens', methods=['POST'])
@login_required
@panel_permission_required('painel36')
def api_painel36_cfg_orig_criar():
    dados = request.get_json() or {}
    nome  = (dados.get('nome') or '').strip()
    if not nome:
        return jsonify({'success': False, 'error': 'Nome e obrigatorio'}), 400
    try:
        with get_db_cursor() as cursor:
            cursor.execute("""
                INSERT INTO padioleiro_origens (nome, ativo, ordem)
                VALUES (%s, TRUE, %s) RETURNING id
            """, (nome, dados.get('ordem', 0)))
            return jsonify({'success': True, 'id': cursor.fetchone()['id']}), 201
    except Exception as e:
        current_app.logger.error(f'Erro criar origem painel36: {e}', exc_info=True)
        return jsonify({'success': False, 'error': 'Erro ao criar origem'}), 500


@painel36_bp.route('/api/paineis/painel36/config/origens/<int:origem_id>', methods=['PUT'])
@login_required
@panel_permission_required('painel36')
def api_painel36_cfg_orig_atualizar(origem_id):
    return _cfg_atualizar(
        'padioleiro_origens', _CAMPOS_ORIGEM,
        request.get_json() or {}, origem_id,
        'Origem atualizada', 'atualizar origem painel36', strip_str=True,
    )

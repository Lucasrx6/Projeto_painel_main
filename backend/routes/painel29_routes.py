# ============================================================
# PAINEL 29 - GESTÃO SENTIR E AGIR
# Hospital Anchieta Ceilândia
# Dashboard, Listagem, Edição (admin), Export Excel
# ============================================================

import os
import io
import traceback
from datetime import datetime, date, timedelta
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from decimal import Decimal
from flask import Blueprint, request, jsonify, send_from_directory, send_file, session, current_app
from psycopg2.extras import RealDictCursor
from backend.database import get_db_connection, release_connection
from backend.middleware.decorators import login_required
from backend.user_management import verificar_permissao_painel

painel29_bp = Blueprint('painel29', __name__)


# ============================================================
# HELPERS
# ============================================================

def _get_ip():
    return request.headers.get('X-Forwarded-For', request.remote_addr)


def _get_usuario():
    return session.get('usuario', 'sistema')


def _is_admin():
    return session.get('is_admin', False)


def _extrair_obs_item_str(descricao_problema):
    if not descricao_problema:
        return None
    if ' | Observacao do item: ' in descricao_problema:
        return descricao_problema.split(' | Observacao do item: ', 1)[1].strip()
    return None


def _registrar_log(cursor, entidade, entidade_id, acao, usuario,
                   campo_alterado=None, valor_anterior=None, valor_novo=None,
                   ip_origem=None):
    cursor.execute("""
        INSERT INTO sentir_agir_log
            (entidade, entidade_id, acao, campo_alterado,
             valor_anterior, valor_novo, usuario, ip_origem)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
    """, (entidade, entidade_id, acao, campo_alterado,
          valor_anterior, valor_novo, usuario, ip_origem))


def serializar_linha(row):
    resultado = {}
    for chave, valor in row.items():
        if isinstance(valor, (datetime, date)):
            resultado[chave] = valor.isoformat()
        elif isinstance(valor, Decimal):
            resultado[chave] = float(valor)
        else:
            resultado[chave] = valor
    return resultado


def _parse_multi_param(param_name):
    raw = request.args.get(param_name, None)
    if not raw or raw.strip() == '':
        return None
    valores = [v.strip() for v in raw.split(',') if v.strip()]
    return valores if valores else None


def _build_common_filters():
    """Constrói filtros compartilhados entre dashboard e dados."""
    condicoes = ["r.status != 'cancelada'"]
    params = []

    # Periodo
    dias = request.args.get('dias', None)
    dt_inicio = request.args.get('dt_inicio', None)
    dt_fim = request.args.get('dt_fim', None)

    if dt_inicio:
        condicoes.append("r.data_ronda >= %s")
        params.append(dt_inicio)
    if dt_fim:
        condicoes.append("r.data_ronda <= %s")
        params.append(dt_fim)
    if dias and not dt_inicio and not dt_fim:
        condicoes.append("r.data_ronda >= CURRENT_DATE - %s * INTERVAL '1 day'")
        params.append(int(dias))

    # Setor
    setores = _parse_multi_param('setor')
    if setores:
        placeholders = ','.join(['%s'] * len(setores))
        condicoes.append("s.nome IN (" + placeholders + ")")
        params.extend(setores)

    # Dupla
    duplas = _parse_multi_param('dupla')
    if duplas:
        placeholders = ','.join(['%s'] * len(duplas))
        condicoes.append("d.id::text IN (" + placeholders + ")")
        params.extend(duplas)

    # Avaliação final
    avaliacoes = _parse_multi_param('avaliacao')
    if avaliacoes:
        placeholders = ','.join(['%s'] * len(avaliacoes))
        condicoes.append("v.avaliacao_final IN (" + placeholders + ")")
        params.extend(avaliacoes)

    # Status da ronda
    status_ronda = _parse_multi_param('status_ronda')
    if status_ronda:
        # Remove o filtro padrão de cancelada e usa o filtro do usuário
        condicoes = [c for c in condicoes if c != "r.status != 'cancelada'"]
        placeholders = ','.join(['%s'] * len(status_ronda))
        condicoes.append("r.status IN (" + placeholders + ")")
        params.extend(status_ronda)

    # Filtro por status de tratativa
    status_trat = request.args.get('status_trat', '').strip()
    if status_trat:
        condicoes.append("""EXISTS (
            SELECT 1 FROM sentir_agir_tratativas t2
            WHERE t2.visita_id = v.id AND t2.status = %s
        )""")
        params.append(status_trat)

    # Busca livre
    busca = request.args.get('busca', '').strip()
    if busca:
        condicoes.append("""(
            v.leito ILIKE %s OR
            v.nr_atendimento ILIKE %s OR
            v.observacoes ILIKE %s OR
            s.nome ILIKE %s OR
            d.nome_visitante_1 ILIKE %s OR
            d.nome_visitante_2 ILIKE %s
        )""")
        termo = '%' + busca + '%'
        params.extend([termo] * 6)

    return condicoes, params


# ============================================================
# ROTAS HTML / ESTÁTICOS
# ============================================================

@painel29_bp.route('/painel/painel29')
@login_required
def painel29():
    usuario_id = session.get('usuario_id')
    is_admin = session.get('is_admin', False)
    if not is_admin:
        if not verificar_permissao_painel(usuario_id, 'painel29'):
            return send_from_directory('frontend', 'acesso-negado.html')
    return send_from_directory('paineis/painel29', 'index.html')


@painel29_bp.route('/paineis/painel29/<path:filename>')
@login_required
def painel29_static(filename):
    return send_from_directory('paineis/painel29', filename)


# ============================================================
# API: DASHBOARD (KPIs)
# ============================================================

@painel29_bp.route('/api/paineis/painel29/dashboard', methods=['GET'])
@login_required
def dashboard():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        condicoes, params = _build_common_filters()
        where = " AND ".join(condicoes) if condicoes else "TRUE"

        sql = """
            SELECT
                COUNT(DISTINCT v.id) AS total_visitas,
                COUNT(DISTINCT r.id) AS total_rondas,
                COUNT(DISTINCT v.leito) AS total_leitos,
                COUNT(DISTINCT r.dupla_id) AS total_duplas_ativas,
                SUM(CASE WHEN v.avaliacao_final = 'critico' THEN 1 ELSE 0 END) AS total_criticos,
                SUM(CASE WHEN v.avaliacao_final = 'atencao' THEN 1 ELSE 0 END) AS total_atencao,
                SUM(CASE WHEN v.avaliacao_final = 'adequado' THEN 1 ELSE 0 END) AS total_adequados,
                (SELECT COUNT(*) FROM sentir_agir_imagens i2
                 JOIN sentir_agir_visitas v2 ON v2.id = i2.visita_id
                 JOIN sentir_agir_rondas r2 ON r2.id = v2.ronda_id
                 WHERE r2.status != 'cancelada') AS total_imagens
            FROM sentir_agir_visitas v
            JOIN sentir_agir_rondas r ON r.id = v.ronda_id
            JOIN sentir_agir_setores s ON s.id = v.setor_id
            JOIN sentir_agir_duplas d ON d.id = r.dupla_id
            WHERE """ + where

        cursor.execute(sql, params)
        stats = cursor.fetchone()

        # KPIs de tratativas (filtradas pelo mesmo período via join)
        cursor.execute("""
            SELECT
                COUNT(*) AS trat_total,
                SUM(CASE WHEN t.status = 'pendente' THEN 1 ELSE 0 END) AS trat_pendentes,
                SUM(CASE WHEN t.status = 'em_tratativa' THEN 1 ELSE 0 END) AS trat_em_tratativa,
                SUM(CASE WHEN t.status = 'regularizado' THEN 1 ELSE 0 END) AS trat_regularizadas,
                SUM(CASE WHEN t.status = 'impossibilitado' THEN 1 ELSE 0 END) AS trat_impossibilitados,
                SUM(CASE WHEN t.status = 'cancelado' THEN 1 ELSE 0 END) AS trat_canceladas
            FROM sentir_agir_tratativas t
            JOIN sentir_agir_visitas v ON v.id = t.visita_id
            JOIN sentir_agir_rondas r ON r.id = v.ronda_id
            JOIN sentir_agir_setores s ON s.id = v.setor_id
            JOIN sentir_agir_duplas d ON d.id = r.dupla_id
            WHERE """ + where, params)
        stats_trat = cursor.fetchone()

        # Top 5 itens mais críticos (ranking)
        cursor.execute("""
            SELECT
                i.descricao AS item,
                c.nome AS categoria,
                COUNT(*) FILTER (WHERE a.resultado = 'critico') AS qtd_critico,
                COUNT(*) FILTER (WHERE a.resultado != 'nao_aplica') AS qtd_total,
                ROUND(
                    100.0 * COUNT(*) FILTER (WHERE a.resultado = 'critico') /
                    NULLIF(COUNT(*) FILTER (WHERE a.resultado != 'nao_aplica'), 0), 1
                ) AS pct_critico
            FROM sentir_agir_avaliacoes a
            JOIN sentir_agir_itens i ON i.id = a.item_id
            JOIN sentir_agir_categorias c ON c.id = i.categoria_id
            JOIN sentir_agir_visitas v ON v.id = a.visita_id
            JOIN sentir_agir_rondas r ON r.id = v.ronda_id
            WHERE r.status != 'cancelada'
            GROUP BY i.descricao, c.nome
            HAVING COUNT(*) FILTER (WHERE a.resultado = 'critico') > 0
            ORDER BY qtd_critico DESC, pct_critico DESC
            LIMIT 5
        """)
        top_criticos = [serializar_linha(r) for r in cursor.fetchall()]

        # KPI: visitas impossibilitadas no período
        cursor.execute("""
            SELECT COUNT(*) AS total_impossibilitadas
            FROM sentir_agir_visitas v
            JOIN sentir_agir_rondas r ON r.id = v.ronda_id
            JOIN sentir_agir_setores s ON s.id = v.setor_id
            JOIN sentir_agir_duplas d ON d.id = r.dupla_id
            WHERE """ + where + " AND v.avaliacao_final = 'impossibilitada'", params)
        stats_impos = cursor.fetchone()

        # KPI: total de pacientes em precaução de contato (global, não filtrado por período)
        cursor.execute("SELECT COUNT(*) AS total_precaucao FROM sentir_agir_precaucao_contato")
        stats_precaucao = cursor.fetchone()

        cursor.close()
        release_connection(conn)

        resultado = serializar_linha(stats) if stats else {}
        resultado['top_criticos'] = top_criticos
        resultado['is_admin'] = _is_admin()
        resultado['total_impossibilitadas'] = int(stats_impos['total_impossibilitadas']) if stats_impos else 0
        resultado['total_precaucao_contato'] = int(stats_precaucao['total_precaucao']) if stats_precaucao else 0

        # Mesclar KPIs de tratativas
        if stats_trat:
            trat = serializar_linha(stats_trat)
            resultado['trat_total'] = trat.get('trat_total', 0) or 0
            resultado['trat_pendentes'] = trat.get('trat_pendentes', 0) or 0
            resultado['trat_em_tratativa'] = trat.get('trat_em_tratativa', 0) or 0
            resultado['trat_regularizadas'] = trat.get('trat_regularizadas', 0) or 0
            resultado['trat_impossibilitados'] = trat.get('trat_impossibilitados', 0) or 0

        return jsonify({'success': True, 'data': resultado})
    except Exception as e:
        current_app.logger.error("Erro no endpoint: %s", e, exc_info=True)
        return jsonify({'success': False, 'error': 'Erro interno do servidor'}), 500


# ============================================================
# API: DADOS (Listagem com filtros)
# ============================================================

@painel29_bp.route('/api/paineis/painel29/dados', methods=['GET'])
@login_required
def dados():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        condicoes, params = _build_common_filters()
        where = " AND ".join(condicoes) if condicoes else "TRUE"

        sql = """
            SELECT
                v.id AS visita_id,
                r.id AS ronda_id,
                r.data_ronda,
                r.status AS status_ronda,
                r.criado_por,
                d.id AS dupla_id,
                d.nome_visitante_1 || ' e ' || d.nome_visitante_2 AS dupla_nome,
                s.nome AS setor_nome,
                s.sigla AS setor_sigla,
                v.leito,
                v.nr_atendimento,
                v.nm_paciente,
                v.avaliacao_final,
                v.observacoes,
                v.criado_em,
                v.status_tratativa,
                (SELECT COUNT(*) FROM sentir_agir_avaliacoes a
                 WHERE a.visita_id = v.id AND a.resultado = 'critico') AS qtd_critico,
                (SELECT COUNT(*) FROM sentir_agir_avaliacoes a
                 WHERE a.visita_id = v.id AND a.resultado = 'atencao') AS qtd_atencao,
                (SELECT COUNT(*) FROM sentir_agir_avaliacoes a
                 WHERE a.visita_id = v.id AND a.resultado = 'adequado') AS qtd_adequado,
                (SELECT COUNT(*) FROM sentir_agir_avaliacoes a
                 WHERE a.visita_id = v.id AND a.resultado = 'nao_aplica') AS qtd_nao_aplica,
                (SELECT COUNT(*) FROM sentir_agir_imagens i
                 WHERE i.visita_id = v.id) AS qtd_imagens,
                (SELECT COUNT(*) FROM sentir_agir_tratativas t
                 WHERE t.visita_id = v.id) AS trat_total,
                (SELECT COUNT(*) FROM sentir_agir_tratativas t
                 WHERE t.visita_id = v.id AND t.status = 'pendente') AS trat_pendentes,
                (SELECT COUNT(*) FROM sentir_agir_tratativas t
                 WHERE t.visita_id = v.id AND t.status = 'em_tratativa') AS trat_em_tratativa,
                (SELECT COUNT(*) FROM sentir_agir_tratativas t
                 WHERE t.visita_id = v.id AND t.status = 'regularizado') AS trat_regularizadas
            FROM sentir_agir_visitas v
            JOIN sentir_agir_rondas r ON r.id = v.ronda_id
            JOIN sentir_agir_setores s ON s.id = v.setor_id
            JOIN sentir_agir_duplas d ON d.id = r.dupla_id
            WHERE """ + where + """
            ORDER BY r.data_ronda DESC, v.criado_em DESC
        """

        cursor.execute(sql, params)
        rows = cursor.fetchall()

        cursor.close()
        release_connection(conn)

        dados_serializados = [serializar_linha(r) for r in rows]

        return jsonify({
            'success': True,
            'data': dados_serializados,
            'total': len(dados_serializados),
            'is_admin': _is_admin()
        })
    except Exception as e:
        current_app.logger.error("Erro no endpoint: %s", e, exc_info=True)
        return jsonify({'success': False, 'error': 'Erro interno do servidor'}), 500


# ============================================================
# API: FILTROS (valores distintos)
# ============================================================

@painel29_bp.route('/api/paineis/painel29/filtros', methods=['GET'])
@login_required
def filtros():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        # Setores
        cursor.execute("""
            SELECT DISTINCT s.nome
            FROM sentir_agir_setores s
            JOIN sentir_agir_visitas v ON v.setor_id = s.id
            ORDER BY s.nome
        """)
        setores = [r['nome'] for r in cursor.fetchall()]

        # Duplas
        cursor.execute("""
            SELECT DISTINCT d.id, d.nome_visitante_1 || ' e ' || d.nome_visitante_2 AS nome
            FROM sentir_agir_duplas d
            JOIN sentir_agir_rondas r ON r.dupla_id = d.id
            ORDER BY nome
        """)
        duplas = cursor.fetchall()

        # Avaliações
        avaliacoes = ['critico', 'atencao', 'adequado']

        # Status ronda
        status_ronda = ['em_andamento', 'concluida', 'cancelada']

        cursor.close()
        release_connection(conn)

        return jsonify({
            'success': True,
            'data': {
                'setores': setores,
                'duplas': [serializar_linha(d) for d in duplas],
                'avaliacoes': avaliacoes,
                'status_ronda': status_ronda
            }
        })
    except Exception as e:
        current_app.logger.error("Erro no endpoint: %s", e, exc_info=True)
        return jsonify({'success': False, 'error': 'Erro interno do servidor'}), 500


# ============================================================
# API: DETALHE DE VISITA (reusa lógica do P28)
# ============================================================

@painel29_bp.route('/api/paineis/painel29/visitas/<int:visita_id>', methods=['GET'])
@login_required
def detalhe_visita(visita_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        cursor.execute("""
            SELECT
                v.id, v.ronda_id, v.leito, v.nr_atendimento,
                v.observacoes, v.avaliacao_final, v.setor_id,
                v.criado_em, v.atualizado_em,
                s.nome AS setor_nome, s.sigla AS setor_sigla,
                r.data_ronda, r.status AS status_ronda,
                d.nome_visitante_1 || ' e ' || d.nome_visitante_2 AS dupla_nome
            FROM sentir_agir_visitas v
            JOIN sentir_agir_setores s ON s.id = v.setor_id
            JOIN sentir_agir_rondas r ON r.id = v.ronda_id
            JOIN sentir_agir_duplas d ON d.id = r.dupla_id
            WHERE v.id = %s
        """, (visita_id,))
        visita = cursor.fetchone()

        if not visita:
            cursor.close()
            release_connection(conn)
            return jsonify({'success': False, 'error': 'Visita nao encontrada'}), 404

        # Avaliações agrupadas por categoria (inclui tratativa_id e obs_item)
        cursor.execute("""
            SELECT
                a.id AS avaliacao_id, a.item_id, a.resultado,
                i.descricao AS item_descricao, i.ordem AS item_ordem,
                c.id AS categoria_id, c.nome AS categoria_nome,
                c.icone AS categoria_icone, c.cor AS categoria_cor,
                c.ordem AS categoria_ordem, c.permite_nao_aplica,
                (SELECT t.id FROM sentir_agir_tratativas t
                 WHERE t.visita_id = a.visita_id AND t.item_id = a.item_id
                   AND t.status NOT IN ('cancelado') ORDER BY t.id LIMIT 1) AS tratativa_id,
                (SELECT t.descricao_problema FROM sentir_agir_tratativas t
                 WHERE t.visita_id = a.visita_id AND t.item_id = a.item_id
                   AND t.status NOT IN ('cancelado') ORDER BY t.id LIMIT 1) AS trat_descricao_problema
            FROM sentir_agir_avaliacoes a
            JOIN sentir_agir_itens i ON i.id = a.item_id
            JOIN sentir_agir_categorias c ON c.id = i.categoria_id
            WHERE a.visita_id = %s
            ORDER BY c.ordem, i.ordem
        """, (visita_id,))
        avaliacoes_raw = cursor.fetchall()

        # Imagens
        cursor.execute("""
            SELECT id, caminho_arquivo, nome_original, descricao, criado_em
            FROM sentir_agir_imagens WHERE visita_id = %s ORDER BY criado_em
        """, (visita_id,))
        imagens = cursor.fetchall()

        # Tratativas desta visita
        cursor.execute("""
            SELECT
                t.id AS tratativa_id,
                t.status,
                t.prioridade,
                t.descricao_problema,
                t.plano_acao,
                t.data_resolucao,
                t.criado_em,
                i.descricao AS item_descricao,
                c.nome AS categoria_nome,
                c.icone AS categoria_icone,
                COALESCE(resp.nome, t.responsavel_nome_manual, 'Sem responsavel') AS responsavel_display
            FROM sentir_agir_tratativas t
            JOIN sentir_agir_itens i ON i.id = t.item_id
            JOIN sentir_agir_categorias c ON c.id = t.categoria_id
            LEFT JOIN sentir_agir_responsaveis resp ON resp.id = t.responsavel_id
            WHERE t.visita_id = %s
            ORDER BY
                CASE t.status
                    WHEN 'pendente' THEN 1
                    WHEN 'em_tratativa' THEN 2
                    WHEN 'regularizado' THEN 3
                    WHEN 'cancelado' THEN 4
                END,
                t.criado_em
        """, (visita_id,))
        tratativas = cursor.fetchall()

        # Log de alterações desta visita
        cursor.execute("""
            SELECT acao, campo_alterado, valor_anterior, valor_novo, usuario, criado_em
            FROM sentir_agir_log
            WHERE entidade = 'visita' AND entidade_id = %s
            ORDER BY criado_em DESC LIMIT 20
        """, (visita_id,))
        historico = cursor.fetchall()

        cursor.close()
        release_connection(conn)

        # Serializar visita
        visita_dict = serializar_linha(visita)

        # Agrupar avaliações
        categorias = []
        cat_atual = None
        for av in avaliacoes_raw:
            if cat_atual is None or cat_atual['id'] != av['categoria_id']:
                cat_atual = {
                    'id': av['categoria_id'],
                    'nome': av['categoria_nome'],
                    'icone': av['categoria_icone'],
                    'cor': av['categoria_cor'],
                    'permite_nao_aplica': av['permite_nao_aplica'],
                    'itens': []
                }
                categorias.append(cat_atual)
            cat_atual['itens'].append({
                'avaliacao_id': av['avaliacao_id'],
                'item_id': av['item_id'],
                'descricao': av['item_descricao'],
                'resultado': av['resultado'],
                'tratativa_id': av['tratativa_id'],
                'obs_item': _extrair_obs_item_str(av.get('trat_descricao_problema'))
            })

        visita_dict['categorias'] = categorias
        visita_dict['imagens'] = [
            dict(serializar_linha(img), url='/api/paineis/painel28/imagens/%d' % img['id'])
            for img in imagens
        ]
        visita_dict['tratativas'] = [serializar_linha(t) for t in tratativas]
        visita_dict['historico'] = [serializar_linha(h) for h in historico]
        visita_dict['is_admin'] = _is_admin()

        return jsonify({'success': True, 'data': visita_dict})
    except Exception as e:
        current_app.logger.error("Erro no endpoint: %s", e, exc_info=True)
        return jsonify({'success': False, 'error': 'Erro interno do servidor'}), 500


# ============================================================
# API: EDITAR VISITA (apenas admin)
# ============================================================

@painel29_bp.route('/api/paineis/painel29/visitas/<int:visita_id>', methods=['PUT'])
@login_required
def editar_visita(visita_id):
    """Edita uma visita. Acessivel a qualquer usuario logado."""
    try:
        dados = request.get_json()
        if not dados:
            return jsonify({'success': False, 'error': 'Dados nao fornecidos'}), 400

        usuario = _get_usuario()
        ip = _get_ip()

        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        # Verificar visita
        cursor.execute("SELECT * FROM sentir_agir_visitas WHERE id = %s", (visita_id,))
        visita = cursor.fetchone()
        if not visita:
            cursor.close()
            release_connection(conn)
            return jsonify({'success': False, 'error': 'Visita nao encontrada'}), 404

        # Verificar bloqueio por tempo
        cursor.execute("SELECT valor FROM sentir_agir_config WHERE chave = 'dias_bloquear_edicao'")
        config_dias = cursor.fetchone()
        if config_dias and config_dias['valor'] != '0':
            dias_bloqueio = int(config_dias['valor'])
            if visita['criado_em']:
                limite = visita['criado_em'] + timedelta(days=dias_bloqueio)
                if datetime.now() > limite:
                    cursor.close()
                    release_connection(conn)
                    return jsonify({
                        'success': False,
                        'error': 'Edicao bloqueada. Prazo de %d dias expirado.' % dias_bloqueio
                    }), 403

        alteracoes = []

        # Atualizar campos básicos
        campos_editaveis = {
            'leito': 'leito',
            'nr_atendimento': 'nr_atendimento',
            'observacoes': 'observacoes',
            'avaliacao_final': 'avaliacao_final',
            'setor_id': 'setor_id'
        }

        sets = []
        set_params = []

        for campo_json, campo_db in campos_editaveis.items():
            if campo_json in dados:
                novo_valor = dados[campo_json]
                antigo_valor = visita.get(campo_db)

                if str(novo_valor) != str(antigo_valor):
                    sets.append(campo_db + " = %s")
                    set_params.append(novo_valor)
                    alteracoes.append({
                        'campo': campo_db,
                        'anterior': str(antigo_valor),
                        'novo': str(novo_valor)
                    })

        if sets:
            sets.append("atualizado_em = NOW()")
            set_params.append(visita_id)
            cursor.execute(
                "UPDATE sentir_agir_visitas SET " + ", ".join(sets) + " WHERE id = %s",
                set_params
            )

        # Atualizar avaliações individuais e obs_item das críticas
        if 'avaliacoes' in dados and isinstance(dados['avaliacoes'], list):
            for av in dados['avaliacoes']:
                av_id = av.get('avaliacao_id')
                novo_resultado = av.get('resultado')
                trat_id = av.get('tratativa_id')
                novo_obs = (av.get('obs_item') or '').strip() or None

                if av_id and novo_resultado:
                    cursor.execute(
                        "SELECT resultado FROM sentir_agir_avaliacoes WHERE id = %s AND visita_id = %s",
                        (av_id, visita_id)
                    )
                    av_atual = cursor.fetchone()
                    if av_atual and av_atual['resultado'] != novo_resultado:
                        cursor.execute(
                            "UPDATE sentir_agir_avaliacoes SET resultado = %s WHERE id = %s",
                            (novo_resultado, av_id)
                        )
                        alteracoes.append({
                            'campo': 'avaliacao_' + str(av_id),
                            'anterior': av_atual['resultado'],
                            'novo': novo_resultado
                        })

                # Atualizar obs_item na tratativa, se fornecido
                if trat_id and 'obs_item' in av:
                    cursor.execute(
                        "SELECT descricao_problema FROM sentir_agir_tratativas WHERE id = %s AND visita_id = %s",
                        (trat_id, visita_id)
                    )
                    trat = cursor.fetchone()
                    if trat:
                        desc_atual = trat['descricao_problema'] or ''
                        obs_atual = _extrair_obs_item_str(desc_atual)
                        if obs_atual != novo_obs:
                            if ' | Observacao do item: ' in desc_atual:
                                base_desc = desc_atual.split(' | Observacao do item: ')[0]
                            elif ' | Observacao da visita: ' in desc_atual:
                                base_desc = desc_atual.split(' | Observacao da visita: ')[0]
                            else:
                                base_desc = desc_atual
                            nova_desc = base_desc + (' | Observacao do item: ' + novo_obs if novo_obs else '')
                            cursor.execute(
                                "UPDATE sentir_agir_tratativas SET descricao_problema = %s WHERE id = %s",
                                (nova_desc, trat_id)
                            )
                            alteracoes.append({
                                'campo': 'obs_critica_trat_' + str(trat_id),
                                'anterior': obs_atual or '',
                                'novo': novo_obs or ''
                            })

        # Registrar logs
        for alt in alteracoes:
            _registrar_log(
                cursor, 'visita', visita_id, 'edicao', usuario,
                campo_alterado=alt['campo'],
                valor_anterior=alt['anterior'],
                valor_novo=alt['novo'],
                ip_origem=ip
            )

        conn.commit()
        cursor.close()
        release_connection(conn)

        return jsonify({
            'success': True,
            'message': 'Visita atualizada. %d campo(s) alterado(s).' % len(alteracoes),
            'alteracoes': len(alteracoes)
        })
    except Exception as e:
        current_app.logger.error("Erro no endpoint: %s", e, exc_info=True)
        return jsonify({'success': False, 'error': 'Erro interno do servidor'}), 500


# ============================================================
# API: ALTERAR STATUS DA RONDA (qualquer usuário logado)
# ============================================================

@painel29_bp.route('/api/paineis/painel29/rondas/<int:ronda_id>/status', methods=['PUT'])
@login_required
def alterar_status_ronda(ronda_id):
    """Altera o status de uma ronda. Acessível a qualquer usuário logado."""
    try:
        dados = request.get_json()
        if not dados:
            return jsonify({'success': False, 'error': 'Dados nao fornecidos'}), 400

        novo_status = dados.get('status', '').strip()
        status_validos = ('em_andamento', 'concluida', 'cancelada')
        if novo_status not in status_validos:
            return jsonify({'success': False, 'error': 'Status invalido. Use: ' + ', '.join(status_validos)}), 400

        usuario = _get_usuario()
        ip = _get_ip()
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        cursor.execute("SELECT id, status FROM sentir_agir_rondas WHERE id = %s", (ronda_id,))
        ronda = cursor.fetchone()
        if not ronda:
            cursor.close()
            release_connection(conn)
            return jsonify({'success': False, 'error': 'Ronda nao encontrada'}), 404

        if ronda['status'] == novo_status:
            cursor.close()
            release_connection(conn)
            return jsonify({'success': True, 'message': 'Ronda ja esta com o status ' + novo_status})

        cursor.execute(
            "UPDATE sentir_agir_rondas SET status = %s, atualizado_em = NOW() WHERE id = %s",
            (novo_status, ronda_id)
        )
        _registrar_log(cursor, 'ronda', ronda_id, 'alteracao_status', usuario,
                       campo_alterado='status',
                       valor_anterior=ronda['status'],
                       valor_novo=novo_status,
                       ip_origem=ip)
        conn.commit()
        cursor.close()
        release_connection(conn)

        labels = {'em_andamento': 'Em andamento', 'concluida': 'Concluida', 'cancelada': 'Cancelada'}
        return jsonify({'success': True, 'message': 'Status alterado para: ' + labels.get(novo_status, novo_status)})
    except Exception as e:
        current_app.logger.error("Erro no endpoint: %s", e, exc_info=True)
        return jsonify({'success': False, 'error': 'Erro interno do servidor'}), 500


# ============================================================
# API: EXPORTAR EXCEL
# ============================================================

@painel29_bp.route('/api/paineis/painel29/exportar', methods=['GET'])
@login_required
def exportar_excel():
    """Exporta dados filtrados para Excel (.xlsx) com 5 abas completas."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        condicoes, params = _build_common_filters()
        where = " AND ".join(condicoes) if condicoes else "TRUE"

        # ── 1. RESUMO / KPIs ─────────────────────────────────────────
        cursor.execute("""
            SELECT
                COUNT(DISTINCT v.id)                                                          AS total_visitas,
                COUNT(DISTINCT r.id)                                                          AS total_rondas,
                COUNT(DISTINCT v.leito)                                                       AS total_leitos,
                COUNT(DISTINCT r.dupla_id)                                                    AS total_duplas,
                SUM(CASE WHEN v.avaliacao_final = 'critico'         THEN 1 ELSE 0 END)       AS total_criticos,
                SUM(CASE WHEN v.avaliacao_final = 'atencao'         THEN 1 ELSE 0 END)       AS total_atencao,
                SUM(CASE WHEN v.avaliacao_final = 'adequado'        THEN 1 ELSE 0 END)       AS total_adequados,
                SUM(CASE WHEN v.avaliacao_final = 'impossibilitada' THEN 1 ELSE 0 END)       AS total_impossibilitadas,
                (SELECT COUNT(*) FROM sentir_agir_tratativas t2
                 JOIN sentir_agir_visitas v2 ON v2.id = t2.visita_id
                 JOIN sentir_agir_rondas  r2 ON r2.id = v2.ronda_id
                 WHERE """ + where + """)                                                     AS trat_total,
                (SELECT COUNT(*) FROM sentir_agir_tratativas t2
                 JOIN sentir_agir_visitas v2 ON v2.id = t2.visita_id
                 JOIN sentir_agir_rondas  r2 ON r2.id = v2.ronda_id
                 WHERE """ + where + """ AND t2.status = 'pendente')                         AS trat_pendentes,
                (SELECT COUNT(*) FROM sentir_agir_tratativas t2
                 JOIN sentir_agir_visitas v2 ON v2.id = t2.visita_id
                 JOIN sentir_agir_rondas  r2 ON r2.id = v2.ronda_id
                 WHERE """ + where + """ AND t2.status = 'em_tratativa')                     AS trat_em_tratativa,
                (SELECT COUNT(*) FROM sentir_agir_tratativas t2
                 JOIN sentir_agir_visitas v2 ON v2.id = t2.visita_id
                 JOIN sentir_agir_rondas  r2 ON r2.id = v2.ronda_id
                 WHERE """ + where + """ AND t2.status = 'regularizado')                     AS trat_regularizadas,
                (SELECT COUNT(*) FROM sentir_agir_tratativas t2
                 JOIN sentir_agir_visitas v2 ON v2.id = t2.visita_id
                 JOIN sentir_agir_rondas  r2 ON r2.id = v2.ronda_id
                 WHERE """ + where + """ AND t2.status = 'impossibilitado')                  AS trat_impossibilitadas,
                MIN(r.data_ronda)                                                             AS data_inicio,
                MAX(r.data_ronda)                                                             AS data_fim
            FROM sentir_agir_visitas v
            JOIN sentir_agir_rondas  r ON r.id = v.ronda_id
            JOIN sentir_agir_setores s ON s.id = v.setor_id
            JOIN sentir_agir_duplas  d ON d.id = r.dupla_id
            WHERE """ + where, params * 5)
        kpis = cursor.fetchone() or {}

        # Top 5 itens críticos
        cursor.execute("""
            SELECT
                c.nome AS categoria,
                i.descricao AS item,
                COUNT(*) AS qtd_critico
            FROM sentir_agir_avaliacoes a
            JOIN sentir_agir_itens     i ON i.id = a.item_id
            JOIN sentir_agir_categorias c ON c.id = i.categoria_id
            JOIN sentir_agir_visitas   v ON v.id = a.visita_id
            JOIN sentir_agir_rondas    r ON r.id = v.ronda_id
            JOIN sentir_agir_setores   s ON s.id = v.setor_id
            JOIN sentir_agir_duplas    d ON d.id = r.dupla_id
            WHERE a.resultado = 'critico' AND """ + where + """
            GROUP BY c.nome, i.descricao
            ORDER BY qtd_critico DESC
            LIMIT 5
        """, params)
        top_criticos = cursor.fetchall()

        # ── 2. VISITAS ────────────────────────────────────────────────
        cursor.execute("""
            SELECT
                r.id                                                                  AS "ID Ronda",
                v.id                                                                  AS "ID Visita",
                r.data_ronda                                                          AS "Data Ronda",
                d.nome_visitante_1 || ' e ' || d.nome_visitante_2                    AS "Dupla",
                s.nome                                                                AS "Setor",
                v.leito                                                               AS "Leito",
                v.nr_atendimento                                                      AS "Nr Atendimento",
                v.nm_paciente                                                         AS "Paciente",
                v.avaliacao_final                                                     AS "Avaliação Final",
                r.status                                                              AS "Status Ronda",
                v.status_tratativa                                                    AS "Status Tratativa",
                v.observacoes                                                         AS "Observações Gerais",
                v.criado_em                                                           AS "Data Registro",
                r.criado_por                                                          AS "Registrado Por",
                (SELECT COUNT(*) FROM sentir_agir_avaliacoes a
                 WHERE a.visita_id = v.id AND a.resultado = 'critico')               AS "Críticos",
                (SELECT COUNT(*) FROM sentir_agir_avaliacoes a
                 WHERE a.visita_id = v.id AND a.resultado = 'atencao')               AS "Atenção",
                (SELECT COUNT(*) FROM sentir_agir_avaliacoes a
                 WHERE a.visita_id = v.id AND a.resultado = 'adequado')              AS "Adequados",
                (SELECT COUNT(*) FROM sentir_agir_avaliacoes a
                 WHERE a.visita_id = v.id AND a.resultado = 'nao_aplica')            AS "Não Aplica",
                (SELECT COUNT(*) FROM sentir_agir_imagens  i WHERE i.visita_id = v.id) AS "Qtd Imagens",
                (SELECT COUNT(*) FROM sentir_agir_tratativas t WHERE t.visita_id = v.id)                           AS "Trat Total",
                (SELECT COUNT(*) FROM sentir_agir_tratativas t WHERE t.visita_id = v.id AND t.status = 'pendente') AS "Trat Pendentes",
                (SELECT COUNT(*) FROM sentir_agir_tratativas t WHERE t.visita_id = v.id AND t.status = 'em_tratativa') AS "Trat Em Tratativa",
                (SELECT COUNT(*) FROM sentir_agir_tratativas t WHERE t.visita_id = v.id AND t.status = 'regularizado') AS "Trat Regularizadas",
                (SELECT COUNT(*) FROM sentir_agir_tratativas t WHERE t.visita_id = v.id AND t.status = 'impossibilitado') AS "Trat Impossibilitadas"
            FROM sentir_agir_visitas v
            JOIN sentir_agir_rondas  r ON r.id = v.ronda_id
            JOIN sentir_agir_setores s ON s.id = v.setor_id
            JOIN sentir_agir_duplas  d ON d.id = r.dupla_id
            WHERE """ + where + """
            ORDER BY r.data_ronda DESC, v.criado_em DESC
        """, params)
        rows_visitas = cursor.fetchall()

        # ── 3. AVALIAÇÕES POR ITEM ────────────────────────────────────
        cursor.execute("""
            SELECT
                r.data_ronda                                                          AS "Data Ronda",
                d.nome_visitante_1 || ' e ' || d.nome_visitante_2                    AS "Dupla",
                s.nome                                                                AS "Setor",
                v.leito                                                               AS "Leito",
                v.nr_atendimento                                                      AS "Nr Atendimento",
                v.nm_paciente                                                         AS "Paciente",
                c.nome                                                                AS "Categoria",
                i.descricao                                                           AS "Item Avaliado",
                a.resultado                                                           AS "Resultado",
                COALESCE(t.descricao_problema, '')                                   AS "Obs / Descrição do Problema",
                COALESCE(t.status, '')                                               AS "Status Tratativa Vinculada"
            FROM sentir_agir_avaliacoes a
            JOIN sentir_agir_itens      i ON i.id = a.item_id
            JOIN sentir_agir_categorias c ON c.id = i.categoria_id
            JOIN sentir_agir_visitas    v ON v.id = a.visita_id
            JOIN sentir_agir_rondas     r ON r.id = v.ronda_id
            JOIN sentir_agir_setores    s ON s.id = v.setor_id
            JOIN sentir_agir_duplas     d ON d.id = r.dupla_id
            LEFT JOIN sentir_agir_tratativas t
                   ON t.visita_id = a.visita_id
                  AND t.item_id   = a.item_id
                  AND t.status NOT IN ('cancelado')
            WHERE """ + where + """
            ORDER BY r.data_ronda DESC, v.id, c.ordem, i.ordem
        """, params)
        rows_aval = cursor.fetchall()

        # ── 4. TRATATIVAS DETALHADAS ──────────────────────────────────
        cursor.execute("""
            SELECT
                r.data_ronda                                                          AS "Data Ronda",
                d.nome_visitante_1 || ' e ' || d.nome_visitante_2                    AS "Dupla",
                s.nome                                                                AS "Setor",
                v.leito                                                               AS "Leito",
                v.nr_atendimento                                                      AS "Nr Atendimento",
                v.nm_paciente                                                         AS "Paciente",
                c.nome                                                                AS "Categoria",
                i.descricao                                                           AS "Item",
                t.status                                                              AS "Status",
                COALESCE(t.prioridade, '')                                            AS "Prioridade",
                t.descricao_problema                                                  AS "Descrição do Problema",
                COALESCE(t.plano_acao, '')                                            AS "Plano de Ação",
                COALESCE(resp.nome, t.responsavel_nome_manual, '')                   AS "Responsável",
                t.criado_em                                                           AS "Aberta Em",
                t.data_inicio_tratativa                                               AS "Início Tratativa",
                t.data_resolucao                                                      AS "Data Resolução",
                COALESCE(t.resolvido_por, '')                                        AS "Resolvido Por",
                COALESCE(t.observacoes_resolucao, '')                                AS "Obs Resolução"
            FROM sentir_agir_tratativas t
            JOIN sentir_agir_visitas    v ON v.id = t.visita_id
            JOIN sentir_agir_rondas     r ON r.id = v.ronda_id
            JOIN sentir_agir_setores    s ON s.id = v.setor_id
            JOIN sentir_agir_duplas     d ON d.id = r.dupla_id
            JOIN sentir_agir_categorias c ON c.id = t.categoria_id
            JOIN sentir_agir_itens      i ON i.id = t.item_id
            LEFT JOIN sentir_agir_responsaveis resp ON resp.id = t.responsavel_id
            WHERE """ + where + """
            ORDER BY
                CASE t.status
                    WHEN 'pendente'       THEN 1
                    WHEN 'em_tratativa'   THEN 2
                    WHEN 'regularizado'   THEN 3
                    WHEN 'impossibilitado' THEN 4
                    ELSE 5
                END,
                r.data_ronda DESC, v.id, t.id
        """, params)
        rows_trat = cursor.fetchall()

        # ── 5. PRECAUÇÃO DE CONTATO ───────────────────────────────────
        cursor.execute("""
            SELECT
                p.nr_atendimento   AS "Nr Atendimento",
                p.nm_paciente      AS "Paciente",
                p.leito            AS "Leito",
                p.marcado_por      AS "Marcado Por",
                p.marcado_em       AS "Marcado Em",
                COALESCE(p.observacao, '') AS "Observação"
            FROM sentir_agir_precaucao_contato p
            ORDER BY p.marcado_em DESC
        """)
        rows_precaucao = cursor.fetchall()

        cursor.close()
        release_connection(conn)

        if not rows_visitas:
            return jsonify({'success': False, 'error': 'Nenhum dado para exportar no período selecionado'}), 404

        # ── Helpers de estilo ─────────────────────────────────────────
        def _fmt_cell(val):
            """Converte valor para tipo nativo do Excel (mantém datas como date/datetime)."""
            if val is None:
                return ''
            if isinstance(val, (datetime, date)):
                return val
            if isinstance(val, Decimal):
                return float(val)
            return val

        COR_HAC       = '1B3A6B'   # azul HAC (cabeçalhos principais)
        COR_SECAO     = '2E86AB'   # azul médio (subtítulos de seção)
        COR_VERDE     = '166534'
        COR_AMARELO   = 'B45309'
        COR_VERMELHO  = '991B1B'
        COR_CINZA     = '6B7280'
        BRANCO        = 'FFFFFF'

        def _fill(hex_color):
            return PatternFill('solid', fgColor=hex_color)

        def _font(bold=False, color=BRANCO, size=10):
            return Font(bold=bold, color=color, size=size, name='Calibri')

        def _borda():
            lado = Side(style='thin', color='D1D5DB')
            return Border(left=lado, right=lado, top=lado, bottom=lado)

        def _centro():
            return Alignment(horizontal='center', vertical='center', wrap_text=True)

        def _esquerda():
            return Alignment(horizontal='left', vertical='center', wrap_text=True)

        def _aplicar_cabecalho(ws, colunas, linha=1,
                               bg=COR_HAC, fg=BRANCO, bold=True, size=10):
            for col_idx, titulo in enumerate(colunas, start=1):
                cel = ws.cell(row=linha, column=col_idx, value=titulo)
                cel.fill      = _fill(bg)
                cel.font      = _font(bold=bold, color=fg, size=size)
                cel.alignment = _centro()
                cel.border    = _borda()

        def _aplicar_linha(ws, dados, linha, zebra=False):
            bg_zebra = 'F0F4FF' if zebra else BRANCO
            for col_idx, val in enumerate(dados, start=1):
                cel = ws.cell(row=linha, column=col_idx, value=_fmt_cell(val))
                cel.fill      = _fill(bg_zebra)
                cel.font      = Font(name='Calibri', size=9, color='111827')
                cel.alignment = _esquerda()
                cel.border    = _borda()
                # Formatar datas
                if isinstance(val, datetime):
                    cel.number_format = 'DD/MM/YYYY HH:MM'
                elif isinstance(val, date):
                    cel.number_format = 'DD/MM/YYYY'

        def _ajustar_colunas(ws, larguras_min=None):
            """Ajusta largura de cada coluna pelo conteúdo, com mínimo configurável."""
            larguras_min = larguras_min or {}
            for col in ws.columns:
                max_len = 0
                col_letra = get_column_letter(col[0].column)
                for cel in col:
                    try:
                        v = str(cel.value) if cel.value is not None else ''
                        max_len = max(max_len, len(v))
                    except Exception:
                        pass
                min_w = larguras_min.get(col_letra, 10)
                ws.column_dimensions[col_letra].width = min(max(max_len + 2, min_w), 60)

        # ── Criar workbook ────────────────────────────────────────────
        wb = Workbook()
        wb.remove(wb.active)   # remove a aba padrão vazia

        periodo_str = ''
        if kpis.get('data_inicio') and kpis.get('data_fim'):
            di = kpis['data_inicio']
            df = kpis['data_fim']
            di_str = di.strftime('%d/%m/%Y') if isinstance(di, date) else str(di)
            df_str = df.strftime('%d/%m/%Y') if isinstance(df, date) else str(df)
            periodo_str = '%s a %s' % (di_str, df_str)

        # ══════════════════════════════════════════════════════════════
        # ABA 1 — RESUMO
        # ══════════════════════════════════════════════════════════════
        ws1 = wb.create_sheet('Resumo')
        ws1.sheet_view.showGridLines = False

        # Título
        ws1.merge_cells('A1:D1')
        t = ws1['A1']
        t.value     = 'GESTÃO SENTIR E AGIR — RESUMO DO PERÍODO'
        t.fill      = _fill(COR_HAC)
        t.font      = _font(bold=True, size=14)
        t.alignment = _centro()
        ws1.row_dimensions[1].height = 32

        ws1.merge_cells('A2:D2')
        p = ws1['A2']
        p.value     = 'Período: %s    |    Gerado em: %s' % (
            periodo_str or 'Todos', datetime.now().strftime('%d/%m/%Y %H:%M'))
        p.fill      = _fill('E8EFF8')
        p.font      = Font(name='Calibri', size=10, color='374151')
        p.alignment = _centro()
        ws1.row_dimensions[2].height = 18

        def _kpi_row(ws, linha, rotulo, valor, cor_bg=COR_HAC):
            cel_r = ws.cell(row=linha, column=1, value=rotulo)
            cel_r.fill      = _fill('F3F4F6')
            cel_r.font      = Font(name='Calibri', size=10, bold=True, color='374151')
            cel_r.alignment = _esquerda()
            cel_r.border    = _borda()

            cel_v = ws.cell(row=linha, column=2, value=valor if valor is not None else 0)
            cel_v.fill      = _fill('FFFFFF')
            cel_v.font      = Font(name='Calibri', size=10, color='111827')
            cel_v.alignment = _centro()
            cel_v.border    = _borda()

        # Bloco Visitas
        ws1.merge_cells('A4:B4')
        s = ws1['A4']
        s.value = 'VISITAS'
        s.fill  = _fill(COR_SECAO); s.font = _font(bold=True, size=11); s.alignment = _centro()
        ws1.row_dimensions[4].height = 20

        _kpi_row(ws1, 5,  'Total de Visitas',        kpis.get('total_visitas', 0))
        _kpi_row(ws1, 6,  'Total de Rondas',          kpis.get('total_rondas', 0))
        _kpi_row(ws1, 7,  'Leitos Distintos',         kpis.get('total_leitos', 0))
        _kpi_row(ws1, 8,  'Duplas Ativas',            kpis.get('total_duplas', 0))
        _kpi_row(ws1, 9,  'Críticos',                 kpis.get('total_criticos', 0))
        _kpi_row(ws1, 10, 'Atenção',                  kpis.get('total_atencao', 0))
        _kpi_row(ws1, 11, 'Adequados',                kpis.get('total_adequados', 0))
        _kpi_row(ws1, 12, 'Impossibilitadas',         kpis.get('total_impossibilitadas', 0))

        # Bloco Tratativas
        ws1.merge_cells('A14:B14')
        s2 = ws1['A14']
        s2.value = 'TRATATIVAS'
        s2.fill  = _fill(COR_SECAO); s2.font = _font(bold=True, size=11); s2.alignment = _centro()
        ws1.row_dimensions[14].height = 20

        _kpi_row(ws1, 15, 'Total de Tratativas',      kpis.get('trat_total', 0))
        _kpi_row(ws1, 16, 'Pendentes',                kpis.get('trat_pendentes', 0))
        _kpi_row(ws1, 17, 'Em Tratativa',             kpis.get('trat_em_tratativa', 0))
        _kpi_row(ws1, 18, 'Regularizadas',            kpis.get('trat_regularizadas', 0))
        _kpi_row(ws1, 19, 'Impossibilitadas',         kpis.get('trat_impossibilitadas', 0))

        # Top 5 críticos
        if top_criticos:
            ws1.merge_cells('A21:B21')
            s3 = ws1['A21']
            s3.value = 'TOP 5 ITENS CRÍTICOS'
            s3.fill  = _fill('991B1B'); s3.font = _font(bold=True, size=11); s3.alignment = _centro()
            ws1.row_dimensions[21].height = 20

            _aplicar_cabecalho(ws1, ['Categoria / Item', 'Qtd Crítico'], linha=22,
                               bg='FEE2E2', fg='991B1B', bold=True)
            for idx, tc in enumerate(top_criticos, start=23):
                _kpi_row(ws1, idx,
                         '%s — %s' % (tc.get('categoria', ''), tc.get('item', '')),
                         tc.get('qtd_critico', 0))

        ws1.column_dimensions['A'].width = 36
        ws1.column_dimensions['B'].width = 18
        ws1.column_dimensions['C'].width = 18
        ws1.column_dimensions['D'].width = 18

        # ══════════════════════════════════════════════════════════════
        # ABA 2 — VISITAS
        # ══════════════════════════════════════════════════════════════
        ws2 = wb.create_sheet('Visitas')
        ws2.sheet_view.showGridLines = False
        ws2.freeze_panes = 'A2'

        if rows_visitas:
            colunas = list(rows_visitas[0].keys())
            _aplicar_cabecalho(ws2, colunas)
            for idx, row in enumerate(rows_visitas, start=2):
                _aplicar_linha(ws2, [row[c] for c in colunas], idx, zebra=(idx % 2 == 0))
            _ajustar_colunas(ws2, {'A': 10, 'B': 10, 'C': 14, 'E': 18,
                                   'F': 16, 'G': 26, 'H': 14})

        # ══════════════════════════════════════════════════════════════
        # ABA 3 — AVALIAÇÕES POR ITEM
        # ══════════════════════════════════════════════════════════════
        ws3 = wb.create_sheet('Avaliações por Item')
        ws3.sheet_view.showGridLines = False
        ws3.freeze_panes = 'A2'

        if rows_aval:
            colunas_a = list(rows_aval[0].keys())
            _aplicar_cabecalho(ws3, colunas_a)
            for idx, row in enumerate(rows_aval, start=2):
                _aplicar_linha(ws3, [row[c] for c in colunas_a], idx, zebra=(idx % 2 == 0))
                # Colorir células de resultado
                resultado = row.get('Resultado', '')
                cel_res = ws3.cell(row=idx, column=colunas_a.index('Resultado') + 1)
                if resultado == 'critico':
                    cel_res.fill = _fill('FEE2E2')
                    cel_res.font = Font(name='Calibri', size=9, color=COR_VERMELHO, bold=True)
                elif resultado == 'atencao':
                    cel_res.fill = _fill('FEF3C7')
                    cel_res.font = Font(name='Calibri', size=9, color=COR_AMARELO, bold=True)
                elif resultado == 'adequado':
                    cel_res.fill = _fill('DCFCE7')
                    cel_res.font = Font(name='Calibri', size=9, color=COR_VERDE, bold=True)
            _ajustar_colunas(ws3, {'A': 14, 'H': 28, 'I': 14, 'J': 40})

        # ══════════════════════════════════════════════════════════════
        # ABA 4 — TRATATIVAS
        # ══════════════════════════════════════════════════════════════
        ws4 = wb.create_sheet('Tratativas')
        ws4.sheet_view.showGridLines = False
        ws4.freeze_panes = 'A2'

        if rows_trat:
            colunas_t = list(rows_trat[0].keys())
            _aplicar_cabecalho(ws4, colunas_t)
            for idx, row in enumerate(rows_trat, start=2):
                _aplicar_linha(ws4, [row[c] for c in colunas_t], idx, zebra=(idx % 2 == 0))
                # Colorir coluna Status
                status = row.get('Status', '')
                cel_st = ws4.cell(row=idx, column=colunas_t.index('Status') + 1)
                if status == 'pendente':
                    cel_st.fill = _fill('FEE2E2')
                    cel_st.font = Font(name='Calibri', size=9, color=COR_VERMELHO, bold=True)
                elif status == 'em_tratativa':
                    cel_st.fill = _fill('FEF3C7')
                    cel_st.font = Font(name='Calibri', size=9, color=COR_AMARELO, bold=True)
                elif status == 'regularizado':
                    cel_st.fill = _fill('DCFCE7')
                    cel_st.font = Font(name='Calibri', size=9, color=COR_VERDE, bold=True)
            _ajustar_colunas(ws4, {'A': 14, 'K': 30, 'L': 36, 'M': 22})

        # ══════════════════════════════════════════════════════════════
        # ABA 5 — PRECAUÇÃO DE CONTATO
        # ══════════════════════════════════════════════════════════════
        ws5 = wb.create_sheet('Precaução de Contato')
        ws5.sheet_view.showGridLines = False
        ws5.freeze_panes = 'A2'

        if rows_precaucao:
            colunas_p = list(rows_precaucao[0].keys())
            _aplicar_cabecalho(ws5, colunas_p)
            for idx, row in enumerate(rows_precaucao, start=2):
                _aplicar_linha(ws5, [row[c] for c in colunas_p], idx, zebra=(idx % 2 == 0))
            _ajustar_colunas(ws5, {'B': 26, 'D': 22, 'F': 36})
        else:
            ws5['A1'].value = 'Nenhum paciente em precaução de contato.'
            ws5['A1'].font  = Font(name='Calibri', size=10, color=COR_CINZA, italic=True)

        # ── Serializar e retornar ─────────────────────────────────────
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        nome_arquivo = 'sentir_agir_%s.xlsx' % datetime.now().strftime('%Y%m%d_%H%M%S')

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nome_arquivo
        )
    except Exception as e:
        current_app.logger.error("Erro no endpoint exportar: %s", e, exc_info=True)
        return jsonify({'success': False, 'error': 'Erro interno do servidor'}), 500


# ============================================================
# API: CONFIGURAÇÕES DO PAINEL
# ============================================================

@painel29_bp.route('/api/paineis/painel29/config', methods=['GET'])
@login_required
def obter_config():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("SELECT chave, valor FROM sentir_agir_config")
        rows = cursor.fetchall()
        cursor.close()
        release_connection(conn)

        config = {}
        for row in rows:
            config[row['chave']] = row['valor']
        config['is_admin'] = _is_admin()

        return jsonify({'success': True, 'data': config})
    except Exception as e:
        current_app.logger.error("Erro no endpoint: %s", e, exc_info=True)
        return jsonify({'success': False, 'error': 'Erro interno do servidor'}), 500


@painel29_bp.route('/api/paineis/painel29/config', methods=['PUT'])
@login_required
def atualizar_config():
    """Atualiza configurações. Apenas admin."""
    if not _is_admin():
        return jsonify({'success': False, 'error': 'Apenas administradores'}), 403

    try:
        dados = request.get_json()
        chave = dados.get('chave', '').strip()
        valor = dados.get('valor', '').strip()

        if not chave:
            return jsonify({'success': False, 'error': 'Chave obrigatoria'}), 400

        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        cursor.execute("SELECT valor FROM sentir_agir_config WHERE chave = %s", (chave,))
        atual = cursor.fetchone()

        if atual:
            cursor.execute(
                "UPDATE sentir_agir_config SET valor = %s, atualizado_em = NOW() WHERE chave = %s",
                (valor, chave)
            )
            _registrar_log(
                cursor, 'config', None, 'edicao', _get_usuario(),
                campo_alterado=chave, valor_anterior=atual['valor'],
                valor_novo=valor, ip_origem=_get_ip()
            )
        else:
            cursor.execute(
                "INSERT INTO sentir_agir_config (chave, valor) VALUES (%s, %s)",
                (chave, valor)
            )

        conn.commit()
        cursor.close()
        release_connection(conn)

        return jsonify({'success': True, 'message': 'Configuracao atualizada'})
    except Exception as e:
        current_app.logger.error("Erro no endpoint: %s", e, exc_info=True)
        return jsonify({'success': False, 'error': 'Erro interno do servidor'}), 500


# ============================================================
# API: PRECAUÇÃO DE CONTATO (gestão via painel29)
# ============================================================

@painel29_bp.route('/api/paineis/painel29/precaucao-contato', methods=['GET'])
@login_required
def listar_precaucao_contato_gestao():
    """Lista todos os pacientes em precaução de contato."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("""
            SELECT nr_atendimento, nm_paciente, leito, marcado_por, marcado_em
            FROM sentir_agir_precaucao_contato
            ORDER BY marcado_em DESC
        """)
        rows = cursor.fetchall()
        cursor.close()
        release_connection(conn)
        resultado = []
        for r in rows:
            item = dict(r)
            if item.get('marcado_em'):
                item['marcado_em'] = item['marcado_em'].isoformat()
            resultado.append(item)
        return jsonify({'success': True, 'data': resultado, 'total': len(resultado)})
    except Exception as e:
        current_app.logger.error("Erro no endpoint: %s", e, exc_info=True)
        return jsonify({'success': False, 'error': 'Erro interno do servidor'}), 500


@painel29_bp.route('/api/paineis/painel29/precaucao-contato/<nr_atendimento>', methods=['DELETE'])
@login_required
def remover_precaucao_contato_gestao(nr_atendimento):
    """Remove a marcação de precaução de contato (disponível no painel de gestão)."""
    if not _is_admin():
        return jsonify({'success': False, 'error': 'Apenas administradores podem remover precaução de contato'}), 403
    nr = str(nr_atendimento).strip()
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "DELETE FROM sentir_agir_precaucao_contato WHERE nr_atendimento = %s", (nr,)
        )
        removido = cursor.rowcount > 0
        conn.commit()
        cursor.close()
        release_connection(conn)
        if not removido:
            return jsonify({'success': False, 'error': 'Paciente não encontrado'}), 404
        return jsonify({'success': True, 'message': 'Precaução de contato removida. Paciente voltou à fila.'})
    except Exception as e:
        current_app.logger.error("Erro no endpoint: %s", e, exc_info=True)
        return jsonify({'success': False, 'error': 'Erro interno do servidor'}), 500